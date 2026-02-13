#!/usr/bin/env python3
"""Tests for the compliance report generator.

Unit tests: YAML parsing, CSV generation, evidence extraction, status determination.
Integration tests: Full pipeline with real audit log data and real taxonomy.
"""

from __future__ import annotations

import csv
import json
import os
import shutil
import tempfile
from pathlib import Path

import openpyxl
import jsonschema

import pytest

# Module under test
from generate import (
    CSV_COLUMNS,
    EVIDENCE_BUNDLE_COLUMNS,
    EVIDENCE_BUNDLE_SCHEMA_VERSION,
    FRAMEWORK_DISPLAY_NAMES,
    FRAMEWORK_REQUIREMENTS,
    GDPR_ART30_ROPA_PATH,
    IMPLEMENTED_MIDDLEWARE,
    STATUS_COLORS,
    CompliancePDF,
    build_evidence_bundle,
    build_evidence_description,
    build_evidence_reference,
    build_implementation_notes,
    build_limitations,
    build_recommendation,
    check_config_exists,
    check_evidence_in_log,
    copy_evidence,
    determine_status,
    generate_rows,
    load_audit_log,
    load_taxonomy,
    main,
    write_csv,
    write_evidence_bundle_csv,
    write_evidence_bundle_json,
    write_pdf,
    write_xlsx,
    _apply_xlsx_formatting,
    _build_control_area_matrix,
    _compute_framework_summary,
    _gather_partial_documented_controls,
    _select_evidence_highlights,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent
TAXONOMY_PATH = SCRIPT_DIR / "control_taxonomy.yaml"
EVIDENCE_SCHEMA_V2_PATH = SCRIPT_DIR / "evidence_schema_v2.json"
TECHNICAL_SCOPE_PATH = (
    PROJECT_ROOT / "docs" / "compliance" / "control-taxonomy-technical-scope.md"
)


@pytest.fixture
def sample_audit_entries():
    """Sample audit log entries matching the real gateway format."""
    return [
        {
            "timestamp": "2026-02-06T07:33:41Z",
            "session_id": "8ef798c7-fb6a-4232-b11a-56d932f55a60",
            "decision_id": "8933b900-a398-42f7-be06-1f7fb8e53acb",
            "trace_id": "ae7773cc-6246-46a3-9602-aec968876f32",
            "spiffe_id": "spiffe://poc.local/agents/mcp-client/dspy-researcher/dev",
            "action": "mcp_request",
            "result": "completed",
            "status_code": 403,
            "security": {"tool_hash_verified": False},
            "prev_hash": "22da7eb451de26f5f17f1f9c335147f1290705128fd0ff7e8bcba622c0f0c92f",
            "bundle_digest": "563d8c8c8cc9e9b5a0d582f90c788ba34f1a583ee0c67dbd5bb6ced8630dc0c8",
            "registry_digest": "03e8fcd3633063c516d8f087f8fcb7e72c2328bf045c3baf9227b0988451185f",
        },
        {
            "timestamp": "2026-02-06T07:33:41Z",
            "session_id": "2c696b0c-bb6f-49ea-a5fe-54668211ca5f",
            "decision_id": "c0bf8f7f-c160-4d21-ba03-5ee03dd206f3",
            "trace_id": "e2e67949-0c41-40f1-ad11-141810868173",
            "spiffe_id": "spiffe://poc.local/agents/mcp-client/dspy-researcher/dev",
            "action": "step_up_gating",
            "result": "gate=step_up allowed=true total_score=4 impact=1",
        },
        {
            "timestamp": "2026-02-06T07:33:42Z",
            "session_id": "f2d0c677-36af-4277-85af-0d4639ac9252",
            "decision_id": "a8916019-5b30-4b44-ab43-f10009123ff9",
            "trace_id": "fe4cd0d6-7513-4173-a4d4-881af3d98dc7",
            "spiffe_id": "spiffe://poc.local/agents/mcp-client/dspy-researcher/dev",
            "action": "mcp_request",
            "result": "completed",
            "status_code": 404,
            "security": {"tool_hash_verified": False},
            "prev_hash": "2f928ca70e61b18f5b47acae72b777512fed08fb5f555b8358d22af4ff3e0f99",
            "bundle_digest": "563d8c8c8cc9e9b5a0d582f90c788ba34f1a583ee0c67dbd5bb6ced8630dc0c8",
            "registry_digest": "03e8fcd3633063c516d8f087f8fcb7e72c2328bf045c3baf9227b0988451185f",
        },
    ]


@pytest.fixture
def sample_control():
    """A single sample control for unit testing."""
    return {
        "id": "GW-AUTH-001",
        "name": "SPIFFE mTLS Authentication",
        "description": "All agent requests authenticated via SPIFFE SVIDs",
        "middleware": "spiffe_auth",
        "step": 3,
        "frameworks": {
            "soc2": ["CC6.1"],
            "iso27001": ["A.9.2.1"],
            "ccpa": [],
            "gdpr": ["Art. 32"],
        },
        "evidence_type": "audit_log",
        "evidence_query": '.spiffe_id != "" and .spiffe_id != null',
    }


@pytest.fixture
def sample_configs():
    """Configuration presence flags for testing."""
    return {
        "opa_policy": True,
        "tool_registry": True,
        "opa_tool_grants": True,
        "spiffe_ids": True,
        "destinations": True,
        "risk_thresholds": True,
        "cosign_key": True,
    }


# ---------------------------------------------------------------------------
# Unit Tests: YAML Parsing
# ---------------------------------------------------------------------------


class TestLoadTaxonomy:
    """Tests for load_taxonomy function."""

    def test_load_real_taxonomy(self):
        """Verify the real taxonomy file loads and contains expected structure."""
        controls = load_taxonomy(TAXONOMY_PATH)
        assert len(controls) > 0, "Taxonomy must contain at least one control"

    def test_taxonomy_has_all_10_control_areas(self):
        """All 10 control areas must be represented."""
        controls = load_taxonomy(TAXONOMY_PATH)
        ids = [c["id"] for c in controls]
        prefixes = {id_.rsplit("-", 1)[0] for id_ in ids}
        expected_prefixes = {
            "GW-AUTH",
            "GW-AUTHZ",
            "GW-DLP",
            "GW-SCAN",
            "GW-AUDIT",
            "GW-SEC",
            "GW-TRANS",
            "GW-AVAIL",
            "GW-SESS",
            "GW-SC",
        }
        assert expected_prefixes.issubset(prefixes), (
            f"Missing control areas: {expected_prefixes - prefixes}"
        )

    def test_taxonomy_control_fields(self):
        """Each control must have required fields."""
        controls = load_taxonomy(TAXONOMY_PATH)
        required_fields = {"id", "name", "description", "frameworks", "evidence_type"}
        for control in controls:
            missing = required_fields - set(control.keys())
            assert not missing, (
                f"Control {control.get('id', '???')} missing fields: {missing}"
            )

    def test_taxonomy_framework_mappings(self):
        """Each control must map to at least one framework."""
        controls = load_taxonomy(TAXONOMY_PATH)
        for control in controls:
            frameworks = control.get("frameworks", {})
            total_mappings = sum(
                len(reqs) for reqs in frameworks.values() if reqs
            )
            assert total_mappings > 0, (
                f"Control {control['id']} has no framework mappings"
            )

    def test_taxonomy_mapping_metadata_fields_present(self):
        """Each technical control exposes required mapping metadata fields."""
        controls = load_taxonomy(TAXONOMY_PATH)
        required = {
            "control_scope",
            "control_family",
            "implementation_tier",
            "evidence_owner",
        }
        for control in controls:
            metadata = control.get("mapping_metadata", {})
            assert required.issubset(metadata.keys()), (
                f"Control {control['id']} missing mapping metadata fields: "
                f"{required - set(metadata.keys())}"
            )
            assert metadata["control_scope"] == "technical"

    def test_taxonomy_controls_have_evidence_paths(self):
        """Every technical control has at least one evidence extraction path."""
        controls = load_taxonomy(TAXONOMY_PATH)
        for control in controls:
            paths = control.get("evidence_paths", [])
            assert isinstance(paths, list)
            assert len(paths) > 0, (
                f"Control {control['id']} has no evidence extraction path"
            )

    def test_taxonomy_includes_expanded_control_set(self):
        """Expanded technical control IDs are present in taxonomy v2.1."""
        controls = load_taxonomy(TAXONOMY_PATH)
        control_ids = {control["id"] for control in controls}
        expected_new_ids = {
            "GW-GOV-001",
            "GW-PRIV-001",
            "GW-PRIV-002",
            "GW-PRIV-003",
        }
        assert expected_new_ids.issubset(control_ids)

    def test_taxonomy_invalid_path_raises(self):
        """Loading from a nonexistent path raises an error."""
        with pytest.raises(FileNotFoundError):
            load_taxonomy(Path("/nonexistent/taxonomy.yaml"))

    def test_taxonomy_empty_raises(self):
        """Loading a YAML with no controls raises ValueError."""
        with tempfile.NamedTemporaryFile(mode="w", suffix=".yaml", delete=False) as f:
            f.write("controls: []\n")
            f.flush()
            with pytest.raises(ValueError, match="No controls found"):
                load_taxonomy(Path(f.name))
            os.unlink(f.name)

    def test_scope_boundary_document_exists_and_is_structured(self):
        """Scope boundary doc distinguishes technical from org/process controls."""
        assert TECHNICAL_SCOPE_PATH.exists()
        text = TECHNICAL_SCOPE_PATH.read_text()
        assert "## In Scope (Technical Controls)" in text
        assert "## Out of Scope (Org/Process Controls)" in text


# ---------------------------------------------------------------------------
# Unit Tests: Audit Log Loading
# ---------------------------------------------------------------------------


class TestLoadAuditLog:
    """Tests for load_audit_log function."""

    def test_load_jsonl(self):
        """Load a valid JSONL file."""
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".jsonl", delete=False
        ) as f:
            f.write(json.dumps({"action": "mcp_request", "status_code": 200}) + "\n")
            f.write(json.dumps({"action": "step_up_gating"}) + "\n")
            f.flush()
            entries = load_audit_log(f.name)
            assert len(entries) == 2
            assert entries[0]["action"] == "mcp_request"
            os.unlink(f.name)

    def test_load_docker_compose_prefixed(self):
        """Handle docker-compose log prefixes gracefully."""
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".jsonl", delete=False
        ) as f:
            f.write(
                'mcp-security-gateway  | 2026/02/06 07:33:41 '
                + json.dumps({"action": "mcp_request", "session_id": "abc"})
                + "\n"
            )
            f.flush()
            entries = load_audit_log(f.name)
            assert len(entries) == 1
            assert entries[0]["session_id"] == "abc"
            os.unlink(f.name)

    def test_load_nonexistent_file(self):
        """Nonexistent file returns empty list."""
        entries = load_audit_log("/nonexistent/audit.jsonl")
        assert entries == []

    def test_load_with_invalid_lines(self):
        """Invalid JSON lines are silently skipped."""
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".jsonl", delete=False
        ) as f:
            f.write("not json at all\n")
            f.write("[AUDIT] Internal tool response: tool=tavily_search\n")
            f.write(json.dumps({"action": "valid"}) + "\n")
            f.flush()
            entries = load_audit_log(f.name)
            assert len(entries) == 1
            assert entries[0]["action"] == "valid"
            os.unlink(f.name)


# ---------------------------------------------------------------------------
# Unit Tests: Evidence Checking
# ---------------------------------------------------------------------------


class TestCheckEvidenceInLog:
    """Tests for check_evidence_in_log and query evaluation."""

    def test_spiffe_id_not_empty(self, sample_audit_entries):
        found, count = check_evidence_in_log(
            sample_audit_entries, '.spiffe_id != "" and .spiffe_id != null'
        )
        assert found is True
        assert count == 3

    def test_action_equals(self, sample_audit_entries):
        found, count = check_evidence_in_log(
            sample_audit_entries, '.action == "mcp_request"'
        )
        assert found is True
        assert count == 2

    def test_action_step_up(self, sample_audit_entries):
        found, count = check_evidence_in_log(
            sample_audit_entries, '.action == "step_up_gating"'
        )
        assert found is True
        assert count == 1

    def test_startswith(self, sample_audit_entries):
        found, count = check_evidence_in_log(
            sample_audit_entries, '.spiffe_id | startswith("spiffe://")'
        )
        assert found is True
        assert count == 3

    def test_contains(self, sample_audit_entries):
        found, count = check_evidence_in_log(
            sample_audit_entries, '.result | contains("total_score")'
        )
        assert found is True
        assert count == 1

    def test_prev_hash_not_null(self, sample_audit_entries):
        found, count = check_evidence_in_log(
            sample_audit_entries, ".prev_hash != null"
        )
        assert found is True
        assert count == 2  # step_up_gating entry has no prev_hash

    def test_status_code_equals(self, sample_audit_entries):
        found, count = check_evidence_in_log(
            sample_audit_entries, ".status_code == 403"
        )
        assert found is True
        assert count == 1

    def test_bundle_digest_not_null(self, sample_audit_entries):
        found, count = check_evidence_in_log(
            sample_audit_entries, ".bundle_digest != null and .registry_digest != null"
        )
        assert found is True
        assert count == 2

    def test_nested_field(self, sample_audit_entries):
        found, count = check_evidence_in_log(
            sample_audit_entries, ".security.tool_hash_verified != null"
        )
        assert found is True
        assert count == 2

    def test_empty_query(self, sample_audit_entries):
        found, count = check_evidence_in_log(sample_audit_entries, None)
        assert found is False
        assert count == 0

    def test_empty_entries(self):
        found, count = check_evidence_in_log([], '.action == "mcp_request"')
        assert found is False
        assert count == 0


# ---------------------------------------------------------------------------
# Unit Tests: Status Determination
# ---------------------------------------------------------------------------


class TestDetermineStatus:
    """Tests for determine_status function."""

    def test_implemented_with_audit_evidence(self, sample_control, sample_configs):
        status = determine_status(sample_control, audit_found=True, configs=sample_configs)
        assert status == "Implemented"

    def test_partial_without_audit_evidence(self, sample_control, sample_configs):
        status = determine_status(sample_control, audit_found=False, configs=sample_configs)
        assert status == "Partial"

    def test_documented_only_for_no_middleware(self, sample_configs):
        control = {
            "id": "GW-SC-001",
            "middleware": None,
            "evidence_type": "configuration",
        }
        status = determine_status(control, audit_found=False, configs=sample_configs)
        assert status == "Documented Only"

    def test_implemented_for_config_evidence(self, sample_configs):
        control = {
            "id": "GW-AUTHZ-002",
            "middleware": "opa",
            "evidence_type": "configuration",
        }
        status = determine_status(control, audit_found=False, configs=sample_configs)
        assert status == "Implemented"

    def test_implemented_for_test_result(self, sample_configs):
        control = {
            "id": "GW-DLP-003",
            "middleware": "dlp",
            "evidence_type": "test_result",
        }
        status = determine_status(control, audit_found=False, configs=sample_configs)
        assert status == "Implemented"


# ---------------------------------------------------------------------------
# Unit Tests: CSV Row Generation
# ---------------------------------------------------------------------------


class TestGenerateRows:
    """Tests for generate_rows function."""

    def test_single_control_produces_correct_rows(
        self, sample_control, sample_audit_entries, sample_configs
    ):
        rows = generate_rows(
            [sample_control], sample_audit_entries, sample_configs, PROJECT_ROOT
        )
        # GW-AUTH-001 maps to soc2:CC6.1, iso27001:A.9.2.1, gdpr:Art.32 (ccpa is empty)
        assert len(rows) == 3
        frameworks_seen = {r["framework"] for r in rows}
        assert frameworks_seen == {"SOC2", "ISO27001", "GDPR"}

    def test_csv_columns_present(
        self, sample_control, sample_audit_entries, sample_configs
    ):
        rows = generate_rows(
            [sample_control], sample_audit_entries, sample_configs, PROJECT_ROOT
        )
        for row in rows:
            for col in CSV_COLUMNS:
                assert col in row, f"Missing column: {col}"

    def test_status_populated(
        self, sample_control, sample_audit_entries, sample_configs
    ):
        rows = generate_rows(
            [sample_control], sample_audit_entries, sample_configs, PROJECT_ROOT
        )
        for row in rows:
            assert row["status"] in ("Implemented", "Partial", "Documented Only")

    def test_framework_requirement_has_description(
        self, sample_control, sample_audit_entries, sample_configs
    ):
        rows = generate_rows(
            [sample_control], sample_audit_entries, sample_configs, PROJECT_ROOT
        )
        for row in rows:
            # Should contain both ID and description
            assert ": " in row["framework_requirement"]

    def test_empty_controls_produces_no_rows(self, sample_audit_entries, sample_configs):
        rows = generate_rows([], sample_audit_entries, sample_configs, PROJECT_ROOT)
        assert len(rows) == 0


# ---------------------------------------------------------------------------
# Unit Tests: CSV Writing
# ---------------------------------------------------------------------------


class TestWriteCSV:
    """Tests for write_csv function."""

    def test_write_and_read_csv(self, sample_control, sample_audit_entries, sample_configs):
        rows = generate_rows(
            [sample_control], sample_audit_entries, sample_configs, PROJECT_ROOT
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "report.csv"
            write_csv(rows, output_path)
            assert output_path.exists()

            with open(output_path, "r") as f:
                reader = csv.DictReader(f)
                read_rows = list(reader)
                assert len(read_rows) == len(rows)
                assert set(reader.fieldnames) == set(CSV_COLUMNS)

    def test_creates_parent_directories(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            nested = Path(tmpdir) / "a" / "b" / "c" / "report.csv"
            write_csv([], nested)
            assert nested.exists()


# ---------------------------------------------------------------------------
# Unit Tests: Evidence Bundle v2
# ---------------------------------------------------------------------------


class TestEvidenceBundleV2:
    """Tests for evidence bundle v2 generation and file outputs."""

    def test_build_evidence_bundle_contains_required_fields(
        self, sample_control, sample_audit_entries, sample_configs
    ):
        rows = generate_rows(
            [sample_control], sample_audit_entries, sample_configs, PROJECT_ROOT
        )
        generated_at = "2026-02-13T00:00:00+00:00"
        bundle = build_evidence_bundle(
            rows,
            [sample_control],
            sample_audit_entries,
            generated_at,
        )

        assert bundle["schema_version"] == EVIDENCE_BUNDLE_SCHEMA_VERSION
        assert bundle["record_count"] == len(bundle["records"])
        assert len(bundle["records"]) == len(rows)

        first = bundle["records"][0]
        required = {
            "control_id",
            "framework",
            "framework_requirement",
            "status",
            "source",
            "timestamp",
            "artifact_reference",
            "control_name",
        }
        assert required.issubset(first.keys())
        assert first["control_id"] == sample_control["id"]
        assert first["source"] == "audit_log"
        assert first["timestamp"] == "2026-02-06T07:33:42Z"

    def test_build_evidence_bundle_validates_against_schema_v2(
        self, sample_control, sample_audit_entries, sample_configs
    ):
        rows = generate_rows(
            [sample_control], sample_audit_entries, sample_configs, PROJECT_ROOT
        )
        bundle = build_evidence_bundle(
            rows,
            [sample_control],
            sample_audit_entries,
            "2026-02-13T00:00:00+00:00",
        )

        with open(EVIDENCE_SCHEMA_V2_PATH, "r") as f:
            schema = json.load(f)
        jsonschema.validate(instance=bundle, schema=schema)

    def test_write_evidence_bundle_json_and_csv(
        self, sample_control, sample_audit_entries, sample_configs
    ):
        rows = generate_rows(
            [sample_control], sample_audit_entries, sample_configs, PROJECT_ROOT
        )
        bundle = build_evidence_bundle(
            rows,
            [sample_control],
            sample_audit_entries,
            "2026-02-13T00:00:00+00:00",
        )

        with tempfile.TemporaryDirectory() as tmpdir:
            out_dir = Path(tmpdir)
            json_path = out_dir / "compliance-evidence.v2.json"
            csv_path = out_dir / "compliance-evidence.v2.csv"

            write_evidence_bundle_json(bundle, json_path)
            write_evidence_bundle_csv(bundle["records"], csv_path)

            assert json_path.exists()
            assert csv_path.exists()

            with open(json_path, "r") as f:
                loaded = json.load(f)
            assert loaded["schema_version"] == EVIDENCE_BUNDLE_SCHEMA_VERSION
            assert loaded["record_count"] == len(loaded["records"])

            with open(csv_path, "r") as f:
                reader = csv.DictReader(f)
                csv_rows = list(reader)
                assert len(csv_rows) == len(bundle["records"])
                assert reader.fieldnames == EVIDENCE_BUNDLE_COLUMNS

    def test_build_evidence_bundle_includes_audit_and_test_sources(
        self, sample_audit_entries
    ):
        controls = load_taxonomy(TAXONOMY_PATH)
        configs = check_config_exists(PROJECT_ROOT)
        rows = generate_rows(controls, sample_audit_entries, configs, PROJECT_ROOT)

        bundle = build_evidence_bundle(
            rows,
            controls,
            sample_audit_entries,
            "2026-02-13T00:00:00+00:00",
        )
        sources = {record["source"] for record in bundle["records"]}
        assert "audit_log" in sources
        assert "test_result" in sources


# ---------------------------------------------------------------------------
# Unit Tests: Helper Functions
# ---------------------------------------------------------------------------


class TestHelperFunctions:
    """Tests for evidence, notes, limitations, recommendation builders."""

    def test_evidence_reference_audit(self, sample_control):
        ref = build_evidence_reference(sample_control, 42, {}, PROJECT_ROOT)
        assert "audit.jsonl" in ref
        assert "42" in ref

    def test_evidence_reference_config(self):
        control = {"evidence_type": "configuration", "middleware": "opa"}
        ref = build_evidence_reference(control, 0, {}, PROJECT_ROOT)
        assert "mcp_policy.rego" in ref

    def test_evidence_reference_test(self):
        control = {"evidence_type": "test_result", "middleware": "dlp"}
        ref = build_evidence_reference(control, 0, {}, PROJECT_ROOT)
        assert "dlp_test.go" in ref

    def test_evidence_description_with_entries(self, sample_control):
        desc = build_evidence_description(sample_control, 10)
        assert "10 audit log entries" in desc

    def test_evidence_description_no_entries(self, sample_control):
        desc = build_evidence_description(sample_control, 0)
        assert "No audit log entries" in desc

    def test_implementation_notes(self, sample_control):
        notes = build_implementation_notes(sample_control)
        assert "spiffe_auth" in notes
        assert "step 3" in notes

    def test_limitations_documented_only(self, sample_control):
        lim = build_limitations(sample_control, "Documented Only")
        assert "design only" in lim

    def test_limitations_implemented_spiffe(self, sample_control):
        lim = build_limitations(sample_control, "Implemented")
        assert "self-signed" in lim

    def test_recommendation_implemented(self, sample_control):
        rec = build_recommendation(sample_control, "Implemented")
        assert "Maintain" in rec

    def test_recommendation_partial(self, sample_control):
        rec = build_recommendation(sample_control, "Partial")
        assert "Complete" in rec


# ---------------------------------------------------------------------------
# Unit Tests: Config Checking
# ---------------------------------------------------------------------------


class TestCheckConfigExists:
    """Tests for check_config_exists function."""

    def test_real_project_root(self):
        configs = check_config_exists(PROJECT_ROOT)
        # These should exist in the real project
        assert configs["opa_policy"] is True
        assert configs["tool_registry"] is True

    def test_nonexistent_root(self):
        configs = check_config_exists(Path("/nonexistent/project"))
        for v in configs.values():
            assert v is False


# ---------------------------------------------------------------------------
# Unit Tests: Framework Requirements Completeness
# ---------------------------------------------------------------------------


class TestFrameworkRequirements:
    """Verify framework requirement mappings are complete."""

    def test_soc2_requirements_present(self):
        assert "CC6.1" in FRAMEWORK_REQUIREMENTS["soc2"]
        assert "CC6.5" in FRAMEWORK_REQUIREMENTS["soc2"]
        assert "CC6.6" in FRAMEWORK_REQUIREMENTS["soc2"]
        assert "CC6.7" in FRAMEWORK_REQUIREMENTS["soc2"]
        assert "CC7.1" in FRAMEWORK_REQUIREMENTS["soc2"]
        assert "CC7.2" in FRAMEWORK_REQUIREMENTS["soc2"]

    def test_iso27001_requirements_present(self):
        expected = [
            "A.8.2.1", "A.9.2.1", "A.9.4.1", "A.10.1.1",
            "A.12.2.1", "A.12.4.1", "A.13.1.1", "A.14.2.7",
        ]
        for req in expected:
            assert req in FRAMEWORK_REQUIREMENTS["iso27001"]

    def test_ccpa_requirements_present(self):
        assert "1798.105" in FRAMEWORK_REQUIREMENTS["ccpa"]
        assert "1798.150" in FRAMEWORK_REQUIREMENTS["ccpa"]

    def test_gdpr_requirements_present(self):
        expected = ["Art. 17", "Art. 25", "Art. 28", "Art. 30", "Art. 32"]
        for req in expected:
            assert req in FRAMEWORK_REQUIREMENTS["gdpr"]


# ---------------------------------------------------------------------------
# Integration Tests: Full Pipeline
# ---------------------------------------------------------------------------


class TestIntegrationFullPipeline:
    """Integration tests that exercise the full generate pipeline."""

    def test_full_pipeline_with_real_taxonomy_no_audit(self):
        """Generate a report from real taxonomy without audit log."""
        controls = load_taxonomy(TAXONOMY_PATH)
        configs = check_config_exists(PROJECT_ROOT)
        audit_entries = []  # No audit log

        rows = generate_rows(controls, audit_entries, configs, PROJECT_ROOT)

        assert len(rows) > 0, "Must produce at least one row"

        # Verify all 10 control areas appear
        control_ids = {r["control_id"] for r in rows}
        control_prefixes = {cid.rsplit("-", 1)[0] for cid in control_ids}
        expected = {
            "GW-AUTH", "GW-AUTHZ", "GW-DLP", "GW-SCAN", "GW-AUDIT",
            "GW-SEC", "GW-TRANS", "GW-AVAIL", "GW-SESS", "GW-SC",
        }
        assert expected.issubset(control_prefixes)

        # Verify all 4 frameworks appear
        frameworks = {r["framework"] for r in rows}
        assert frameworks == {"SOC2", "ISO27001", "CCPA", "GDPR"}

        # Verify all CSV columns present
        for row in rows:
            for col in CSV_COLUMNS:
                assert col in row

    def test_full_pipeline_with_sample_audit(self, sample_audit_entries):
        """Generate a report with sample audit entries (simulating E2E output)."""
        controls = load_taxonomy(TAXONOMY_PATH)
        configs = check_config_exists(PROJECT_ROOT)

        rows = generate_rows(controls, sample_audit_entries, configs, PROJECT_ROOT)
        assert len(rows) > 0

        # Controls with audit_log evidence and matching queries should be Implemented
        auth_rows = [r for r in rows if r["control_id"] == "GW-AUTH-001"]
        assert len(auth_rows) > 0
        for r in auth_rows:
            assert r["status"] == "Implemented"

    def test_full_pipeline_csv_output(self, sample_audit_entries):
        """Full pipeline writes a valid CSV file."""
        controls = load_taxonomy(TAXONOMY_PATH)
        configs = check_config_exists(PROJECT_ROOT)
        rows = generate_rows(controls, sample_audit_entries, configs, PROJECT_ROOT)

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "compliance-report.csv"
            write_csv(rows, output_path)

            # Read back and verify
            with open(output_path, "r") as f:
                reader = csv.DictReader(f)
                read_rows = list(reader)

            assert len(read_rows) == len(rows)
            assert set(reader.fieldnames) == set(CSV_COLUMNS)

            # Verify specific framework requirement values
            soc2_rows = [r for r in read_rows if r["framework"] == "SOC2"]
            assert len(soc2_rows) > 0
            iso_rows = [r for r in read_rows if r["framework"] == "ISO27001"]
            assert len(iso_rows) > 0
            ccpa_rows = [r for r in read_rows if r["framework"] == "CCPA"]
            assert len(ccpa_rows) > 0
            gdpr_rows = [r for r in read_rows if r["framework"] == "GDPR"]
            assert len(gdpr_rows) > 0

    def test_full_pipeline_with_real_audit_log(self):
        """Integration test with the real E2E audit log if available."""
        real_log = PROJECT_ROOT / "tests" / "e2e" / "gateway-audit-logs.log"
        if not real_log.exists():
            pytest.skip("Real audit log not available (run E2E suite first)")

        controls = load_taxonomy(TAXONOMY_PATH)
        configs = check_config_exists(PROJECT_ROOT)
        audit_entries = load_audit_log(str(real_log))

        assert len(audit_entries) > 0, "Real audit log should have entries"

        rows = generate_rows(controls, audit_entries, configs, PROJECT_ROOT)
        assert len(rows) > 0

        # With real audit data, audit-backed controls should be Implemented
        audit_controls = [
            r for r in rows
            if r["evidence_type"] == "audit_log" and r["status"] == "Implemented"
        ]
        assert len(audit_controls) > 0, (
            "With real audit data, some audit-based controls should be Implemented"
        )

    def test_cli_main_with_output(self, sample_audit_entries):
        """Test the main() CLI function end-to-end."""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Create a sample audit log file
            audit_path = Path(tmpdir) / "audit.jsonl"
            with open(audit_path, "w") as f:
                for entry in sample_audit_entries:
                    f.write(json.dumps(entry) + "\n")

            output_dir = Path(tmpdir) / "output"

            exit_code = main([
                "--audit-log", str(audit_path),
                "--output-dir", str(output_dir),
                "--project-root", str(PROJECT_ROOT),
            ])

            assert exit_code == 0
            csv_path = output_dir / "compliance-report.csv"
            evidence_json_path = output_dir / "compliance-evidence.v2.json"
            evidence_csv_path = output_dir / "compliance-evidence.v2.csv"
            assert csv_path.exists()
            assert evidence_json_path.exists()
            assert evidence_csv_path.exists()

            with open(csv_path, "r") as f:
                reader = csv.DictReader(f)
                rows = list(reader)
            assert len(rows) > 0

    def test_cli_main_no_audit_log(self):
        """CLI works even without audit log file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_dir = Path(tmpdir) / "output"
            exit_code = main([
                "--audit-log", "/nonexistent/audit.jsonl",
                "--output-dir", str(output_dir),
                "--project-root", str(PROJECT_ROOT),
            ])
            assert exit_code == 0
            csv_path = output_dir / "compliance-report.csv"
            assert csv_path.exists()

    def test_expanded_taxonomy_controls_emit_rows_without_format_break(
        self, sample_audit_entries
    ):
        """Expanded taxonomy controls appear in output while CSV format stays stable."""
        controls = load_taxonomy(TAXONOMY_PATH)
        configs = check_config_exists(PROJECT_ROOT)
        rows = generate_rows(controls, sample_audit_entries, configs, PROJECT_ROOT)
        row_ids = {row["control_id"] for row in rows}
        assert {"GW-GOV-001", "GW-PRIV-001", "GW-PRIV-002", "GW-PRIV-003"}.issubset(
            row_ids
        )
        for row in rows:
            assert set(row.keys()) == set(CSV_COLUMNS)


# ---------------------------------------------------------------------------
# Integration Tests: Evidence Cross-Referencing
# ---------------------------------------------------------------------------


class TestEvidenceCrossReferencing:
    """Verify evidence references point to real files/resources."""

    def test_opa_policy_referenced_controls(self):
        """Controls referencing OPA policy should point to existing file."""
        controls = load_taxonomy(TAXONOMY_PATH)
        configs = check_config_exists(PROJECT_ROOT)
        rows = generate_rows(controls, [], configs, PROJECT_ROOT)

        opa_rows = [
            r for r in rows
            if "mcp_policy.rego" in r["evidence_reference"]
        ]
        assert len(opa_rows) > 0
        # The file must actually exist
        assert (PROJECT_ROOT / "config" / "opa" / "mcp_policy.rego").exists()

    def test_tool_registry_referenced_controls(self):
        """Controls using tool_registry middleware should reference existing files."""
        controls = load_taxonomy(TAXONOMY_PATH)

        # GW-SC-003 uses tool_registry middleware with audit_log evidence.
        # When audit data is present, it references audit.jsonl; when absent,
        # the test_result fallback references the test file.  Verify both
        # the middleware source and config file exist in the project.
        assert (PROJECT_ROOT / "config" / "tool-registry.yaml").exists()
        assert (
            PROJECT_ROOT / "internal" / "gateway" / "middleware" / "tool_registry.go"
        ).exists()

        # With audit entries matching the query, the reference includes audit.jsonl
        sample_entries = [
            {"security": {"tool_hash_verified": True}, "action": "mcp_request"}
        ]
        configs = check_config_exists(PROJECT_ROOT)
        rows = generate_rows(controls, sample_entries, configs, PROJECT_ROOT)
        sc003_rows = [r for r in rows if r["control_id"] == "GW-SC-003"]
        assert len(sc003_rows) > 0
        assert "audit.jsonl" in sc003_rows[0]["evidence_reference"]

    def test_middleware_test_files_exist(self):
        """Controls referencing test files should have corresponding middleware tests."""
        controls = load_taxonomy(TAXONOMY_PATH)
        for control in controls:
            if control.get("evidence_type") == "test_result" and control.get("middleware"):
                middleware = control["middleware"]
                # Check that either the test file or the middleware file exists
                middleware_file = (
                    PROJECT_ROOT
                    / "internal"
                    / "gateway"
                    / "middleware"
                    / f"{middleware}.go"
                )
                if middleware_file.exists():
                    test_file = (
                        PROJECT_ROOT
                        / "internal"
                        / "gateway"
                        / "middleware"
                        / f"{middleware}_test.go"
                    )
                    assert test_file.exists(), (
                        f"Middleware {middleware} has no test file at {test_file}"
                    )


# ---------------------------------------------------------------------------
# Unit Tests: Framework Summary Computation
# ---------------------------------------------------------------------------


class TestComputeFrameworkSummary:
    """Tests for _compute_framework_summary."""

    def test_summary_counts(self):
        rows = [
            {"framework": "SOC2", "status": "Implemented"},
            {"framework": "SOC2", "status": "Implemented"},
            {"framework": "SOC2", "status": "Partial"},
            {"framework": "GDPR", "status": "Documented Only"},
        ]
        summary = _compute_framework_summary(rows)
        assert summary["SOC2"]["total"] == 3
        assert summary["SOC2"]["Implemented"] == 2
        assert summary["SOC2"]["Partial"] == 1
        assert summary["SOC2"]["pct_implemented"] == 67  # 2/3 = 66.6... rounds to 67
        assert summary["GDPR"]["total"] == 1
        assert summary["GDPR"]["Documented Only"] == 1
        assert summary["GDPR"]["pct_implemented"] == 0

    def test_empty_rows(self):
        summary = _compute_framework_summary([])
        assert summary == {}

    def test_all_implemented(self):
        rows = [
            {"framework": "CCPA", "status": "Implemented"},
            {"framework": "CCPA", "status": "Implemented"},
        ]
        summary = _compute_framework_summary(rows)
        assert summary["CCPA"]["pct_implemented"] == 100


# ---------------------------------------------------------------------------
# Unit Tests: XLSX Formatting
# ---------------------------------------------------------------------------


class TestXLSXFormatting:
    """Tests for XLSX generation functions."""

    def test_write_xlsx_creates_file(
        self, sample_control, sample_audit_entries, sample_configs
    ):
        """XLSX file is created with correct path."""
        rows = generate_rows(
            [sample_control], sample_audit_entries, sample_configs, PROJECT_ROOT
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.xlsx"
            write_xlsx(rows, path)
            assert path.exists()
            assert path.stat().st_size > 0

    def test_xlsx_has_summary_sheet(
        self, sample_control, sample_audit_entries, sample_configs
    ):
        """XLSX has a Summary sheet."""
        rows = generate_rows(
            [sample_control], sample_audit_entries, sample_configs, PROJECT_ROOT
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.xlsx"
            write_xlsx(rows, path)
            wb = openpyxl.load_workbook(str(path))
            assert "Summary" in wb.sheetnames
            wb.close()

    def test_xlsx_has_framework_sheets(
        self, sample_control, sample_audit_entries, sample_configs
    ):
        """XLSX has one sheet per framework present in data."""
        rows = generate_rows(
            [sample_control], sample_audit_entries, sample_configs, PROJECT_ROOT
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.xlsx"
            write_xlsx(rows, path)
            wb = openpyxl.load_workbook(str(path))
            # sample_control maps to SOC2, ISO27001, GDPR
            assert "SOC 2" in wb.sheetnames
            assert "ISO 27001" in wb.sheetnames
            assert "GDPR" in wb.sheetnames
            wb.close()

    def test_xlsx_framework_sheet_has_headers(
        self, sample_control, sample_audit_entries, sample_configs
    ):
        """Framework sheets have CSV_COLUMNS as headers."""
        rows = generate_rows(
            [sample_control], sample_audit_entries, sample_configs, PROJECT_ROOT
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.xlsx"
            write_xlsx(rows, path)
            wb = openpyxl.load_workbook(str(path))
            ws = wb["SOC 2"]
            header_vals = [cell.value for cell in ws[1]]
            assert header_vals == CSV_COLUMNS
            wb.close()

    def test_xlsx_conditional_formatting_colors(
        self, sample_control, sample_audit_entries, sample_configs
    ):
        """Rows with 'Implemented' status get green fill."""
        rows = generate_rows(
            [sample_control], sample_audit_entries, sample_configs, PROJECT_ROOT
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.xlsx"
            write_xlsx(rows, path)
            wb = openpyxl.load_workbook(str(path))
            ws = wb["SOC 2"]
            # Row 2 should be the first data row -- sample_control is Implemented
            status_col_idx = CSV_COLUMNS.index("status") + 1
            status_cell = ws.cell(row=2, column=status_col_idx)
            assert status_cell.value == "Implemented"
            # Check fill color on data cell (any cell in the row)
            first_cell = ws.cell(row=2, column=1)
            # The fill should be the green color
            assert first_cell.fill.start_color.rgb is not None
            fill_hex = first_cell.fill.start_color.rgb
            # openpyxl may prefix with FF for alpha
            assert fill_hex.endswith("C6EFCE") or fill_hex == "00C6EFCE"
            wb.close()

    def test_xlsx_summary_contains_framework_percentages(
        self, sample_control, sample_audit_entries, sample_configs
    ):
        """Summary sheet contains percentage values for frameworks."""
        rows = generate_rows(
            [sample_control], sample_audit_entries, sample_configs, PROJECT_ROOT
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.xlsx"
            write_xlsx(rows, path)
            wb = openpyxl.load_workbook(str(path))
            ws = wb["Summary"]
            # Gather all cell values
            all_vals = []
            for row in ws.iter_rows(values_only=True):
                all_vals.extend([str(v) for v in row if v is not None])
            # Should contain percentage strings
            pct_found = any("%" in v for v in all_vals)
            assert pct_found, f"No percentage found in Summary sheet. Values: {all_vals}"
            wb.close()

    def test_xlsx_summary_has_report_date(
        self, sample_control, sample_audit_entries, sample_configs
    ):
        """Summary sheet contains the report date."""
        from datetime import date as dt_date

        rows = generate_rows(
            [sample_control], sample_audit_entries, sample_configs, PROJECT_ROOT
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.xlsx"
            write_xlsx(rows, path)
            wb = openpyxl.load_workbook(str(path))
            ws = wb["Summary"]
            all_vals = []
            for row in ws.iter_rows(values_only=True):
                all_vals.extend([str(v) for v in row if v is not None])
            today = dt_date.today().isoformat()
            assert any(today in v for v in all_vals), (
                f"Report date {today} not found in Summary. Values: {all_vals}"
            )
            wb.close()

    def test_xlsx_bold_headers(
        self, sample_control, sample_audit_entries, sample_configs
    ):
        """Header row in framework sheets should be bold."""
        rows = generate_rows(
            [sample_control], sample_audit_entries, sample_configs, PROJECT_ROOT
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.xlsx"
            write_xlsx(rows, path)
            wb = openpyxl.load_workbook(str(path))
            ws = wb["SOC 2"]
            for cell in ws[1]:
                assert cell.font.bold is True, (
                    f"Header cell '{cell.value}' is not bold"
                )
            wb.close()

    def test_xlsx_data_row_count_matches_csv(
        self, sample_control, sample_audit_entries, sample_configs
    ):
        """Number of data rows in XLSX framework sheets matches CSV rows."""
        rows = generate_rows(
            [sample_control], sample_audit_entries, sample_configs, PROJECT_ROOT
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.xlsx"
            write_xlsx(rows, path)
            wb = openpyxl.load_workbook(str(path))
            total_xlsx_rows = 0
            for fw_code, display in FRAMEWORK_DISPLAY_NAMES.items():
                if display in wb.sheetnames:
                    ws = wb[display]
                    # Subtract header row
                    total_xlsx_rows += ws.max_row - 1
            assert total_xlsx_rows == len(rows), (
                f"XLSX has {total_xlsx_rows} data rows but CSV has {len(rows)}"
            )
            wb.close()


# ---------------------------------------------------------------------------
# Unit Tests: PDF Generation
# ---------------------------------------------------------------------------


class TestPDFGeneration:
    """Tests for PDF generation functions."""

    def test_write_pdf_creates_file(
        self, sample_control, sample_audit_entries, sample_configs
    ):
        """PDF file is created."""
        rows = generate_rows(
            [sample_control], sample_audit_entries, sample_configs, PROJECT_ROOT
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "summary.pdf"
            num_pages = write_pdf(rows, sample_audit_entries, path)
            assert path.exists()
            assert path.stat().st_size > 0

    def test_pdf_has_4_pages(
        self, sample_control, sample_audit_entries, sample_configs
    ):
        """PDF must have exactly 4 pages."""
        rows = generate_rows(
            [sample_control], sample_audit_entries, sample_configs, PROJECT_ROOT
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "summary.pdf"
            num_pages = write_pdf(rows, sample_audit_entries, path)
            assert num_pages == 4, f"PDF has {num_pages} pages, expected 4"

    def test_pdf_contains_header_text(
        self, sample_control, sample_audit_entries, sample_configs
    ):
        """PDF file should be a valid PDF (starts with %PDF)."""
        rows = generate_rows(
            [sample_control], sample_audit_entries, sample_configs, PROJECT_ROOT
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "summary.pdf"
            write_pdf(rows, sample_audit_entries, path)
            with open(path, "rb") as f:
                header = f.read(5)
            assert header == b"%PDF-", "File is not a valid PDF"

    def test_pdf_with_no_audit_entries(
        self, sample_control, sample_configs
    ):
        """PDF generates correctly even without audit entries."""
        rows = generate_rows(
            [sample_control], [], sample_configs, PROJECT_ROOT
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "summary.pdf"
            num_pages = write_pdf(rows, [], path)
            assert path.exists()
            assert num_pages == 4


# ---------------------------------------------------------------------------
# Unit Tests: Control Area Matrix
# ---------------------------------------------------------------------------


class TestControlAreaMatrix:
    """Tests for _build_control_area_matrix."""

    def test_matrix_structure(self):
        rows = [
            {"control_id": "GW-AUTH-001", "framework": "SOC2", "status": "Implemented"},
            {"control_id": "GW-AUTH-001", "framework": "GDPR", "status": "Partial"},
            {"control_id": "GW-DLP-001", "framework": "SOC2", "status": "Documented Only"},
        ]
        matrix = _build_control_area_matrix(rows)
        assert "GW-AUTH" in matrix
        assert "GW-DLP" in matrix
        assert matrix["GW-AUTH"]["SOC2"] == "Implemented"
        assert matrix["GW-AUTH"]["GDPR"] == "Partial"
        assert matrix["GW-DLP"]["SOC2"] == "Documented Only"

    def test_best_status_wins(self):
        """When multiple controls in same area map to same framework, best status wins."""
        rows = [
            {"control_id": "GW-AUTH-001", "framework": "SOC2", "status": "Partial"},
            {"control_id": "GW-AUTH-002", "framework": "SOC2", "status": "Implemented"},
        ]
        matrix = _build_control_area_matrix(rows)
        assert matrix["GW-AUTH"]["SOC2"] == "Implemented"


# ---------------------------------------------------------------------------
# Unit Tests: Evidence Highlights Selection
# ---------------------------------------------------------------------------


class TestEvidenceHighlights:
    """Tests for _select_evidence_highlights."""

    def test_selects_diverse_actions(self, sample_audit_entries):
        highlights = _select_evidence_highlights(sample_audit_entries)
        actions = {h.get("action") for h in highlights}
        # Sample has mcp_request and step_up_gating
        assert "mcp_request" in actions
        assert "step_up_gating" in actions

    def test_redacts_uuids(self, sample_audit_entries):
        highlights = _select_evidence_highlights(sample_audit_entries)
        for h in highlights:
            if "session_id" in h:
                assert len(h["session_id"]) < 40, (
                    f"session_id not truncated: {h['session_id']}"
                )
                assert h["session_id"].endswith("...")

    def test_max_entries_respected(self, sample_audit_entries):
        highlights = _select_evidence_highlights(sample_audit_entries, max_entries=2)
        assert len(highlights) <= 2

    def test_empty_entries(self):
        highlights = _select_evidence_highlights([])
        assert highlights == []


# ---------------------------------------------------------------------------
# Unit Tests: Partial/Documented Only Gathering
# ---------------------------------------------------------------------------


class TestGatherIncompleteControls:
    """Tests for _gather_partial_documented_controls."""

    def test_filters_implemented(self):
        rows = [
            {"control_id": "GW-A-001", "control_name": "A", "status": "Implemented",
             "recommendation": "ok"},
            {"control_id": "GW-B-001", "control_name": "B", "status": "Partial",
             "recommendation": "fix"},
            {"control_id": "GW-C-001", "control_name": "C", "status": "Documented Only",
             "recommendation": "impl"},
        ]
        result = _gather_partial_documented_controls(rows)
        ids = {r["control_id"] for r in result}
        assert "GW-A-001" not in ids
        assert "GW-B-001" in ids
        assert "GW-C-001" in ids

    def test_deduplicates(self):
        rows = [
            {"control_id": "GW-B-001", "control_name": "B", "status": "Partial",
             "recommendation": "fix", "framework": "SOC2"},
            {"control_id": "GW-B-001", "control_name": "B", "status": "Partial",
             "recommendation": "fix", "framework": "GDPR"},
        ]
        result = _gather_partial_documented_controls(rows)
        assert len(result) == 1


# ---------------------------------------------------------------------------
# Unit Tests: Evidence Copying
# ---------------------------------------------------------------------------


class TestCopyEvidence:
    """Tests for copy_evidence function."""

    def test_copies_audit_excerpt(self, sample_audit_entries):
        """Audit log excerpt is copied to evidence directory."""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Create a sample audit log
            audit_path = Path(tmpdir) / "audit.jsonl"
            with open(audit_path, "w") as f:
                for entry in sample_audit_entries:
                    f.write(json.dumps(entry) + "\n")

            output_dir = Path(tmpdir) / "output"
            copied = copy_evidence(PROJECT_ROOT, str(audit_path), output_dir)

            assert "evidence/audit-log-excerpt.jsonl" in copied
            excerpt = output_dir / "evidence" / "audit-log-excerpt.jsonl"
            assert excerpt.exists()
            with open(excerpt) as f:
                lines = f.readlines()
            assert len(lines) == len(sample_audit_entries)

    def test_copies_policy_configs(self):
        """Policy config files are copied when they exist."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_dir = Path(tmpdir) / "output"
            copied = copy_evidence(PROJECT_ROOT, "/nonexistent", output_dir)

            # OPA policy and tool registry should exist in the real project
            assert "evidence/policy-configs/opa-policy.rego" in copied
            assert "evidence/policy-configs/tool-registry.yaml" in copied
            assert (output_dir / "evidence" / "policy-configs" / "opa-policy.rego").exists()
            assert (output_dir / "evidence" / "policy-configs" / "tool-registry.yaml").exists()

    def test_copies_risk_thresholds(self):
        """Risk thresholds config is copied."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_dir = Path(tmpdir) / "output"
            copied = copy_evidence(PROJECT_ROOT, "/nonexistent", output_dir)
            assert "evidence/policy-configs/risk-thresholds.yaml" in copied

    def test_no_audit_log_still_copies_configs(self):
        """When no audit log exists, configs are still copied."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_dir = Path(tmpdir) / "output"
            copied = copy_evidence(PROJECT_ROOT, "/nonexistent/audit.jsonl", output_dir)
            # Should still have policy configs
            config_files = [f for f in copied if "policy-configs" in f]
            assert len(config_files) > 0

    def test_audit_excerpt_limited_to_50_lines(self):
        """Audit excerpt is limited to first 50 lines."""
        with tempfile.TemporaryDirectory() as tmpdir:
            audit_path = Path(tmpdir) / "big-audit.jsonl"
            with open(audit_path, "w") as f:
                for i in range(100):
                    f.write(json.dumps({"line": i}) + "\n")

            output_dir = Path(tmpdir) / "output"
            copy_evidence(PROJECT_ROOT, str(audit_path), output_dir)

            excerpt = output_dir / "evidence" / "audit-log-excerpt.jsonl"
            with open(excerpt) as f:
                lines = f.readlines()
            assert len(lines) == 50

    def test_e2e_test_results_copied(self):
        """E2E test results file is copied if it exists."""
        real_log = PROJECT_ROOT / "tests" / "e2e" / "gateway-audit-logs.log"
        if not real_log.exists():
            pytest.skip("Real E2E audit log not available")

        with tempfile.TemporaryDirectory() as tmpdir:
            output_dir = Path(tmpdir) / "output"
            copied = copy_evidence(PROJECT_ROOT, "/nonexistent", output_dir)
            assert "evidence/e2e-test-results.txt" in copied


# ---------------------------------------------------------------------------
# Integration Tests: Full Pipeline with All Outputs
# ---------------------------------------------------------------------------


class TestIntegrationAllOutputs:
    """Integration tests generating all three outputs (CSV, XLSX, PDF)."""

    def test_all_three_outputs_from_real_taxonomy(self, sample_audit_entries):
        """Generate CSV, XLSX, and PDF from real taxonomy and sample audit data."""
        controls = load_taxonomy(TAXONOMY_PATH)
        configs = check_config_exists(PROJECT_ROOT)
        rows = generate_rows(controls, sample_audit_entries, configs, PROJECT_ROOT)

        with tempfile.TemporaryDirectory() as tmpdir:
            output_dir = Path(tmpdir) / "report"

            # CSV
            csv_path = output_dir / "compliance-report.csv"
            write_csv(rows, csv_path)

            # XLSX
            xlsx_path = output_dir / "compliance-report.xlsx"
            write_xlsx(rows, xlsx_path)

            # PDF
            pdf_path = output_dir / "compliance-summary.pdf"
            num_pages = write_pdf(rows, sample_audit_entries, pdf_path)

            # Verify all exist
            assert csv_path.exists()
            assert xlsx_path.exists()
            assert pdf_path.exists()

            # PDF has 4 pages
            assert num_pages == 4

    def test_csv_and_xlsx_data_consistency(self, sample_audit_entries):
        """CSV and XLSX contain consistent data."""
        controls = load_taxonomy(TAXONOMY_PATH)
        configs = check_config_exists(PROJECT_ROOT)
        rows = generate_rows(controls, sample_audit_entries, configs, PROJECT_ROOT)

        with tempfile.TemporaryDirectory() as tmpdir:
            csv_path = Path(tmpdir) / "report.csv"
            xlsx_path = Path(tmpdir) / "report.xlsx"

            write_csv(rows, csv_path)
            write_xlsx(rows, xlsx_path)

            # Read CSV
            with open(csv_path) as f:
                reader = csv.DictReader(f)
                csv_rows = list(reader)

            # Read XLSX (all framework sheets)
            wb = openpyxl.load_workbook(str(xlsx_path))
            xlsx_rows: list[dict[str, str]] = []
            for fw_display in ["SOC 2", "ISO 27001", "CCPA", "GDPR"]:
                if fw_display in wb.sheetnames:
                    ws = wb[fw_display]
                    headers = [cell.value for cell in ws[1]]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        xlsx_rows.append(dict(zip(headers, row)))
            wb.close()

            # Same total count
            assert len(csv_rows) == len(xlsx_rows), (
                f"CSV has {len(csv_rows)} rows, XLSX has {len(xlsx_rows)}"
            )

            # Same control IDs
            csv_ids = sorted(r["control_id"] for r in csv_rows)
            xlsx_ids = sorted(r["control_id"] for r in xlsx_rows)
            assert csv_ids == xlsx_ids

            # Same statuses
            csv_statuses = sorted(r["status"] for r in csv_rows)
            xlsx_statuses = sorted(r["status"] for r in xlsx_rows)
            assert csv_statuses == xlsx_statuses

    def test_xlsx_has_all_4_framework_sheets_plus_summary(self):
        """Full taxonomy XLSX has Summary + 4 framework sheets."""
        controls = load_taxonomy(TAXONOMY_PATH)
        configs = check_config_exists(PROJECT_ROOT)
        rows = generate_rows(controls, [], configs, PROJECT_ROOT)

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.xlsx"
            write_xlsx(rows, path)
            wb = openpyxl.load_workbook(str(path))
            assert "Summary" in wb.sheetnames
            assert "SOC 2" in wb.sheetnames
            assert "ISO 27001" in wb.sheetnames
            assert "CCPA" in wb.sheetnames
            assert "GDPR" in wb.sheetnames
            assert len(wb.sheetnames) == 5
            wb.close()

    def test_pdf_with_real_audit_log(self):
        """Integration test with real E2E audit log for PDF generation."""
        real_log = PROJECT_ROOT / "tests" / "e2e" / "gateway-audit-logs.log"
        if not real_log.exists():
            pytest.skip("Real audit log not available (run E2E suite first)")

        controls = load_taxonomy(TAXONOMY_PATH)
        configs = check_config_exists(PROJECT_ROOT)
        audit_entries = load_audit_log(str(real_log))
        rows = generate_rows(controls, audit_entries, configs, PROJECT_ROOT)

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "summary.pdf"
            num_pages = write_pdf(rows, audit_entries, path)
            assert path.exists()
            assert num_pages == 4
            assert path.stat().st_size > 1000  # PDF should be non-trivial

    def test_cli_main_produces_all_outputs(self, sample_audit_entries):
        """CLI main() produces CSV, XLSX, and PDF."""
        with tempfile.TemporaryDirectory() as tmpdir:
            audit_path = Path(tmpdir) / "audit.jsonl"
            with open(audit_path, "w") as f:
                for entry in sample_audit_entries:
                    f.write(json.dumps(entry) + "\n")

            output_dir = Path(tmpdir) / "output"

            exit_code = main([
                "--audit-log", str(audit_path),
                "--output-dir", str(output_dir),
                "--project-root", str(PROJECT_ROOT),
            ])

            assert exit_code == 0
            assert (output_dir / "compliance-report.csv").exists()
            assert (output_dir / "compliance-report.xlsx").exists()
            assert (output_dir / "compliance-summary.pdf").exists()
            assert (output_dir / "compliance-evidence.v2.json").exists()
            assert (output_dir / "compliance-evidence.v2.csv").exists()

    def test_cli_main_produces_evidence_directory(self, sample_audit_entries):
        """CLI main() copies evidence files."""
        with tempfile.TemporaryDirectory() as tmpdir:
            audit_path = Path(tmpdir) / "audit.jsonl"
            with open(audit_path, "w") as f:
                for entry in sample_audit_entries:
                    f.write(json.dumps(entry) + "\n")

            output_dir = Path(tmpdir) / "output"

            exit_code = main([
                "--audit-log", str(audit_path),
                "--output-dir", str(output_dir),
                "--project-root", str(PROJECT_ROOT),
            ])

            assert exit_code == 0
            evidence_dir = output_dir / "evidence"
            assert evidence_dir.exists()
            assert (evidence_dir / "audit-log-excerpt.jsonl").exists()
            assert (evidence_dir / "policy-configs").exists()

    def test_full_pipeline_with_real_audit_all_outputs(self):
        """Full integration: real audit log -> all three outputs + evidence."""
        real_log = PROJECT_ROOT / "tests" / "e2e" / "gateway-audit-logs.log"
        if not real_log.exists():
            pytest.skip("Real audit log not available (run E2E suite first)")

        with tempfile.TemporaryDirectory() as tmpdir:
            output_dir = Path(tmpdir) / "output"

            exit_code = main([
                "--audit-log", str(real_log),
                "--output-dir", str(output_dir),
                "--project-root", str(PROJECT_ROOT),
            ])

            assert exit_code == 0

            # All three outputs
            assert (output_dir / "compliance-report.csv").exists()
            assert (output_dir / "compliance-report.xlsx").exists()
            assert (output_dir / "compliance-summary.pdf").exists()

            # Evidence
            assert (output_dir / "evidence" / "audit-log-excerpt.jsonl").exists()
            assert (output_dir / "evidence" / "policy-configs" / "opa-policy.rego").exists()

            # CSV and XLSX have consistent data
            with open(output_dir / "compliance-report.csv") as f:
                csv_row_count = sum(1 for _ in csv.reader(f)) - 1  # subtract header

            wb = openpyxl.load_workbook(str(output_dir / "compliance-report.xlsx"))
            xlsx_row_count = 0
            for fw_display in ["SOC 2", "ISO 27001", "CCPA", "GDPR"]:
                if fw_display in wb.sheetnames:
                    xlsx_row_count += wb[fw_display].max_row - 1
            wb.close()

            assert csv_row_count == xlsx_row_count, (
                f"CSV {csv_row_count} rows != XLSX {xlsx_row_count} rows"
            )


# ---------------------------------------------------------------------------
# Validation Tests: GDPR Article 30 ROPA Document (RFA-w4m)
# ---------------------------------------------------------------------------


ROPA_PATH = PROJECT_ROOT / "docs" / "compliance" / "gdpr-article-30-ropa.md"


class TestGDPRArticle30ROPA:
    """Validate that the ROPA document contains all required sections per Article 30(1)."""

    def test_ropa_file_exists(self):
        """ROPA document must exist at the expected path."""
        assert ROPA_PATH.exists(), (
            f"ROPA document not found at {ROPA_PATH}"
        )

    def test_ropa_is_non_empty(self):
        """ROPA document must not be empty."""
        content = ROPA_PATH.read_text()
        assert len(content) > 100, "ROPA document is too short to be complete"

    def test_ropa_section_1_controller_processor(self):
        """Section 1: Controller and Processor identification."""
        content = ROPA_PATH.read_text()
        assert "## 1." in content, "Missing Section 1 heading"
        assert "Controller" in content, "Missing Controller identification"
        assert "Processor" in content, "Missing Processor identification"
        assert "Data Controller" in content, "Missing Data Controller subsection"
        assert "Data Processor" in content, "Missing Data Processor subsection"
        # Controller is the deploying org, processor is the gateway
        assert "deploying organization" in content.lower(), (
            "Should identify deploying organization as controller"
        )
        assert "MCP Security Gateway" in content, (
            "Should identify MCP Security Gateway as processor"
        )

    def test_ropa_section_2_data_subject_categories(self):
        """Section 2: Categories of data subjects."""
        content = ROPA_PATH.read_text()
        assert "## 2." in content, "Missing Section 2 heading"
        assert "Data Subjects" in content, "Missing data subjects heading"
        assert "SPIFFE ID" in content, "Must mention SPIFFE IDs as identifiers"
        assert "pseudonymous" in content.lower(), (
            "Must note SPIFFE IDs are pseudonymous identifiers"
        )

    def test_ropa_section_3_processing_categories(self):
        """Section 3: Categories of processing activities."""
        content = ROPA_PATH.read_text()
        assert "## 3." in content, "Missing Section 3 heading"
        # All required processing categories
        assert "Session" in content, "Missing session tracking processing"
        assert "Tool Action" in content or "tool action" in content.lower(), (
            "Missing tool action recording processing"
        )
        assert "Risk Score" in content or "risk score" in content.lower(), (
            "Missing risk score computation processing"
        )
        assert "Exfiltration" in content or "exfiltration" in content.lower(), (
            "Missing exfiltration detection processing"
        )
        assert "Audit" in content, "Missing audit logging processing"
        assert "Rate Limit" in content or "rate limit" in content.lower(), (
            "Missing rate limiting processing"
        )

    def test_ropa_section_4_purpose(self):
        """Section 4: Purpose of processing."""
        content = ROPA_PATH.read_text()
        assert "## 4." in content, "Missing Section 4 heading"
        assert "Purpose" in content, "Missing purpose heading"
        # Required purposes
        assert "security monitoring" in content.lower(), (
            "Missing security monitoring purpose"
        )
        assert "audit trail" in content.lower(), (
            "Missing audit trail purpose"
        )
        assert "rate limiting" in content.lower(), (
            "Missing rate limiting purpose"
        )

    def test_ropa_section_5_retention(self):
        """Section 5: Data retention periods."""
        content = ROPA_PATH.read_text()
        assert "## 5." in content, "Missing Section 5 heading"
        assert "Retention" in content, "Missing retention heading"
        # Specific retention values
        assert "3600" in content, "Missing session TTL of 3600s"
        assert "120" in content, "Missing rate limit TTL of 120s"
        assert "indefinite" in content.lower() or "Indefinite" in content, (
            "Missing indefinite retention for audit logs"
        )

    def test_ropa_section_6_technical_measures(self):
        """Section 6: Technical and organizational measures (Art. 32 cross-ref)."""
        content = ROPA_PATH.read_text()
        assert "## 6." in content, "Missing Section 6 heading"
        assert "Article 32" in content or "Art. 32" in content, (
            "Must reference Article 32"
        )
        # Required technical measures
        assert "Encryption in Transit" in content or "encryption in transit" in content.lower(), (
            "Missing encryption in transit measure"
        )
        assert "mTLS" in content, "Missing mTLS reference"
        assert "SPIRE" in content, "Missing SPIRE reference for mTLS"
        assert "SVID" in content, "Missing SVID reference"
        assert "Right-to-Deletion" in content or "right-to-deletion" in content.lower() or "gdpr-delete" in content, (
            "Missing right-to-deletion measure"
        )
        assert "Access Control" in content or "access control" in content.lower(), (
            "Missing access control measure"
        )
        # OTel Collector exception per story learnings
        assert "OTel Collector" in content or "OpenTelemetry Collector" in content, (
            "Must document OTel Collector mTLS exception"
        )

    def test_ropa_section_7_third_country_transfers(self):
        """Section 7: Transfers to third countries."""
        content = ROPA_PATH.read_text()
        assert "## 7." in content, "Missing Section 7 heading"
        assert "Third Countr" in content or "third countr" in content.lower(), (
            "Missing third country transfers heading"
        )
        # Default: no transfers
        assert "no personal data is transferred" in content.lower() or "none" in content.lower(), (
            "Must state no transfers by default"
        )
        # Groq API risk
        assert "Groq" in content, "Must document Groq API as transfer risk"
        assert "United States" in content or "US-based" in content, (
            "Must document Groq API is US-based"
        )

    def test_ropa_all_7_sections_present(self):
        """All 7 required ROPA sections must be present as numbered headings."""
        content = ROPA_PATH.read_text()
        for i in range(1, 8):
            assert f"## {i}." in content, (
                f"Missing required ROPA section {i}"
            )

    def test_ropa_keydb_session_processing_documented(self):
        """KeyDB session data processing must be explicitly documented."""
        content = ROPA_PATH.read_text()
        assert "KeyDB" in content, "KeyDB must be mentioned as storage backend"
        assert "session:" in content, (
            "KeyDB key format session:{spiffe_id}:{session_id} should be documented"
        )

    def test_ropa_audit_log_processing_documented(self):
        """Audit log processing must be explicitly documented."""
        content = ROPA_PATH.read_text()
        assert "JSONL" in content, "Audit log format (JSONL) must be documented"
        assert "append-only" in content.lower(), (
            "Audit log append-only nature must be documented"
        )

    def test_ropa_groq_third_country_risk_documented(self):
        """Deep scan to Groq API must be documented as a third-country data transfer risk."""
        content = ROPA_PATH.read_text()
        # Find the third country section specifically
        section_7_start = content.find("## 7.")
        assert section_7_start >= 0, "Section 7 not found"
        section_7_content = content[section_7_start:]
        assert "Groq" in section_7_content, (
            "Groq API must be documented in the third-country transfers section"
        )
        assert "deep scan" in section_7_content.lower(), (
            "Deep scan feature must be mentioned in transfer risk context"
        )


# ---------------------------------------------------------------------------
# Integration Test: make gdpr-ropa Target (RFA-w4m)
# ---------------------------------------------------------------------------


class TestMakeGDPRROPA:
    """Integration tests for the 'make gdpr-ropa' Makefile target."""

    def test_gdpr_ropa_target_exists_in_makefile(self):
        """The gdpr-ropa target must exist in the Makefile."""
        makefile_path = PROJECT_ROOT / "Makefile"
        content = makefile_path.read_text()
        assert "gdpr-ropa:" in content, (
            "gdpr-ropa target not found in Makefile"
        )

    def test_gdpr_ropa_output_non_empty(self):
        """'make gdpr-ropa' must produce non-empty output."""
        import subprocess

        result = subprocess.run(
            ["make", "gdpr-ropa"],
            capture_output=True,
            text=True,
            cwd=str(PROJECT_ROOT),
            timeout=30,
        )
        assert result.returncode == 0, (
            f"'make gdpr-ropa' failed with return code {result.returncode}: {result.stderr}"
        )
        assert len(result.stdout) > 100, (
            f"'make gdpr-ropa' output is too short ({len(result.stdout)} chars)"
        )

    def test_gdpr_ropa_output_contains_required_headings(self):
        """'make gdpr-ropa' output must contain all 7 required ROPA section headings."""
        import subprocess

        result = subprocess.run(
            ["make", "gdpr-ropa"],
            capture_output=True,
            text=True,
            cwd=str(PROJECT_ROOT),
            timeout=30,
        )
        assert result.returncode == 0
        output = result.stdout

        # All 7 section headings
        required_headings = [
            "Controller and Processor",
            "Categories of Data Subjects",
            "Categories of Processing",
            "Purpose of Processing",
            "Data Retention",
            "Technical and Organizational Measures",
            "Transfers to Third Countries",
        ]
        for heading in required_headings:
            assert heading in output, (
                f"Required ROPA heading '{heading}' not found in 'make gdpr-ropa' output"
            )

    def test_gdpr_ropa_output_contains_gdpr_reference(self):
        """Output must reference GDPR Article 30."""
        import subprocess

        result = subprocess.run(
            ["make", "gdpr-ropa"],
            capture_output=True,
            text=True,
            cwd=str(PROJECT_ROOT),
            timeout=30,
        )
        assert result.returncode == 0
        assert "Article 30" in result.stdout, (
            "Output must reference GDPR Article 30"
        )


# ---------------------------------------------------------------------------
# Cross-Reference Test: Compliance Report Generator and ROPA (RFA-w4m)
# ---------------------------------------------------------------------------


class TestROPACrossReference:
    """Verify the compliance report generator cross-references the ROPA document."""

    def test_gdpr_art30_ropa_path_constant(self):
        """GDPR_ART30_ROPA_PATH constant must point to existing file."""
        ropa_full_path = PROJECT_ROOT / GDPR_ART30_ROPA_PATH
        assert ropa_full_path.exists(), (
            f"GDPR_ART30_ROPA_PATH points to {ropa_full_path} which does not exist"
        )

    def test_art30_rows_reference_ropa(self):
        """Compliance rows mapped to GDPR Art. 30 must reference the ROPA document."""
        controls = load_taxonomy(TAXONOMY_PATH)
        configs = check_config_exists(PROJECT_ROOT)
        rows = generate_rows(controls, [], configs, PROJECT_ROOT)

        art30_rows = [
            r for r in rows
            if r["framework"] == "GDPR" and "Art. 30" in r["framework_requirement"]
        ]
        assert len(art30_rows) > 0, (
            "No GDPR Art. 30 rows found in compliance report"
        )
        for row in art30_rows:
            assert GDPR_ART30_ROPA_PATH in row["implementation_notes"], (
                f"Control {row['control_id']} mapped to Art. 30 does not reference "
                f"ROPA document in implementation_notes. Got: {row['implementation_notes']}"
            )

    def test_art30_rows_with_audit_data_reference_ropa(self, sample_audit_entries):
        """Art. 30 rows still reference ROPA even when audit data is present."""
        controls = load_taxonomy(TAXONOMY_PATH)
        configs = check_config_exists(PROJECT_ROOT)
        rows = generate_rows(controls, sample_audit_entries, configs, PROJECT_ROOT)

        art30_rows = [
            r for r in rows
            if r["framework"] == "GDPR" and "Art. 30" in r["framework_requirement"]
        ]
        assert len(art30_rows) > 0
        for row in art30_rows:
            assert GDPR_ART30_ROPA_PATH in row["implementation_notes"], (
                f"Control {row['control_id']} Art. 30 row missing ROPA reference "
                f"when audit data present. Got: {row['implementation_notes']}"
            )

    def test_non_art30_rows_do_not_reference_ropa(self):
        """Rows NOT mapped to Art. 30 should NOT reference the ROPA document."""
        controls = load_taxonomy(TAXONOMY_PATH)
        configs = check_config_exists(PROJECT_ROOT)
        rows = generate_rows(controls, [], configs, PROJECT_ROOT)

        non_art30_rows = [
            r for r in rows
            if not ("Art. 30" in r.get("framework_requirement", ""))
        ]
        assert len(non_art30_rows) > 0
        for row in non_art30_rows:
            assert GDPR_ART30_ROPA_PATH not in row["implementation_notes"], (
                f"Control {row['control_id']} ({row['framework_requirement']}) "
                f"should NOT reference ROPA. Got: {row['implementation_notes']}"
            )
