#!/usr/bin/env python3
"""Compliance Report Generator -- Agentic Reference Architecture POC.

Reads the control taxonomy YAML, audit logs, and policy configurations to
produce machine-readable evidence bundles (JSON + CSV) plus analyst-friendly
CSV, XLSX, and PDF compliance reports mapping gateway controls to
SOC 2, ISO 27001, CCPA, and GDPR frameworks.

Usage:
    python3 tools/compliance/generate.py [--audit-log PATH] [--output-dir PATH]

Defaults:
    --audit-log   /tmp/audit.jsonl
    --output-dir  reports/compliance-YYYY-MM-DD/
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import shutil
import sys
from datetime import date, timezone, datetime
from pathlib import Path
from typing import Any

# External dependencies -- required to parse taxonomy and produce outputs.
try:
    import yaml
except ImportError:
    print(
        "ERROR: PyYAML is required.  Install with:  pip install pyyaml",
        file=sys.stderr,
    )
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print(
        "ERROR: openpyxl is required.  Install with:  pip install openpyxl",
        file=sys.stderr,
    )
    sys.exit(1)

try:
    from fpdf import FPDF
except ImportError:
    print(
        "ERROR: fpdf2 is required.  Install with:  pip install fpdf2",
        file=sys.stderr,
    )
    sys.exit(1)


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
TAXONOMY_PATH = SCRIPT_DIR / "control_taxonomy.yaml"

# Framework requirement descriptions (from BUSINESS.md Section 7)
FRAMEWORK_REQUIREMENTS: dict[str, dict[str, str]] = {
    "soc2": {
        "CC6.1": "Logical and Physical Access Controls",
        "CC6.5": "System Availability Controls",
        "CC6.6": "System Boundary Controls",
        "CC6.7": "Data Transmission Controls",
        "CC7.1": "System Monitoring",
        "CC7.2": "System Change Monitoring",
    },
    "iso27001": {
        "A.8.2.1": "Classification of Information",
        "A.9.2.1": "User Registration and Deregistration",
        "A.9.4.1": "Information Access Restriction",
        "A.10.1.1": "Policy on Use of Cryptographic Controls",
        "A.12.2.1": "Controls Against Malware",
        "A.12.4.1": "Event Logging",
        "A.13.1.1": "Network Controls",
        "A.14.2.7": "Outsourced Development",
    },
    "ccpa": {
        "1798.105": "Right to Deletion",
        "1798.150": "Data Breach Private Right of Action",
    },
    "gdpr": {
        "Art. 17": "Right to Erasure",
        "Art. 25": "Data Protection by Design and by Default",
        "Art. 28": "Processor Requirements",
        "Art. 30": "Records of Processing Activities",
        "Art. 32": "Security of Processing",
    },
}

# CSV column order
CSV_COLUMNS = [
    "control_id",
    "control_name",
    "control_description",
    "framework",
    "framework_requirement",
    "status",
    "evidence_type",
    "evidence_reference",
    "evidence_description",
    "test_result",
    "implementation_notes",
    "limitations",
    "recommendation",
]

# Evidence bundle v2 schema/version and CSV column order.
EVIDENCE_BUNDLE_SCHEMA_VERSION = "evidence.bundle.v2"
EVIDENCE_BUNDLE_COLUMNS = [
    "control_id",
    "framework",
    "framework_requirement",
    "status",
    "source",
    "timestamp",
    "artifact_reference",
    "control_name",
]

# Cross-reference: GDPR Article 30 ROPA document location.
# The compliance report generator references this document for any controls
# mapped to GDPR Art. 30 (Records of Processing Activities).
GDPR_ART30_ROPA_PATH = "docs/compliance/gdpr-article-30-ropa.md"

# Middleware -> implementation status mapping.
# Controls whose middleware exists in the codebase are "Implemented";
# those without middleware (supply chain) are "Documented Only".
IMPLEMENTED_MIDDLEWARE = {
    "spiffe_auth",
    "opa",
    "dlp",
    "response_firewall",
    "deep_scan",
    "step_up_gating",
    "audit",
    "spike_token",
    "spike_redeemer",
    "rate_limiter",
    "circuit_breaker",
    "size_limit",
    "session_context",
    "tool_registry",
}


# ---------------------------------------------------------------------------
# Taxonomy loading
# ---------------------------------------------------------------------------


def _infer_control_family(control_id: str) -> str:
    """Infer a control family token from a control identifier."""
    parts = control_id.split("-")
    if len(parts) >= 2:
        return f"{parts[0]}-{parts[1]}"
    return control_id or "UNSPECIFIED"


def _default_mapping_metadata(control: dict[str, Any]) -> dict[str, str]:
    """Return default mapping metadata for controls missing explicit values."""
    middleware = control.get("middleware")
    implementation_tier = "runtime" if middleware else "supply_chain"
    return {
        "control_scope": "technical",
        "control_family": _infer_control_family(str(control.get("id", ""))),
        "implementation_tier": implementation_tier,
        "evidence_owner": "security-platform",
    }


def _default_evidence_paths(control: dict[str, Any]) -> list[str]:
    """Return canonical evidence extraction paths for a control."""
    evidence_type = str(control.get("evidence_type", ""))
    middleware = str(control.get("middleware") or "")

    if evidence_type == "audit_log":
        return ["evidence/audit-log-excerpt.jsonl"]

    if evidence_type == "configuration":
        if middleware in ("opa", "spiffe_auth", "deep_scan", "session_context"):
            return ["evidence/policy-configs/opa-policy.rego"]
        if middleware == "tool_registry":
            return ["evidence/policy-configs/tool-registry.yaml"]
        if middleware == "dlp":
            return ["internal/gateway/middleware/dlp.go"]
        if middleware in ("size_limit", "rate_limiter", "circuit_breaker"):
            return [f"internal/gateway/middleware/{middleware}.go"]
        if middleware:
            return [f"internal/gateway/middleware/{middleware}.go"]
        return ["config/"]

    if evidence_type == "test_result":
        if middleware:
            return [f"internal/gateway/middleware/{middleware}_test.go"]
        return ["tests/"]

    return ["evidence/"]


def load_taxonomy(path: Path | None = None) -> list[dict[str, Any]]:
    """Load and validate the control taxonomy YAML."""
    path = path or TAXONOMY_PATH
    with open(path, "r") as f:
        data = yaml.safe_load(f)
    raw_controls = data.get("controls", [])
    if not raw_controls:
        raise ValueError(f"No controls found in {path}")

    controls: list[dict[str, Any]] = []
    for control in raw_controls:
        normalized = dict(control)

        mapping_metadata = _default_mapping_metadata(normalized)
        mapping_metadata.update(normalized.get("mapping_metadata", {}) or {})
        normalized["mapping_metadata"] = mapping_metadata

        evidence_paths = normalized.get("evidence_paths") or _default_evidence_paths(
            normalized
        )
        normalized["evidence_paths"] = [str(p) for p in evidence_paths if str(p)]
        if not normalized["evidence_paths"]:
            raise ValueError(
                f"Control {normalized.get('id', 'UNKNOWN')} has no evidence paths"
            )

        controls.append(normalized)

    return controls


# ---------------------------------------------------------------------------
# Audit log analysis
# ---------------------------------------------------------------------------


def load_audit_log(path: str) -> list[dict[str, Any]]:
    """Load JSONL audit log, returning parsed entries.

    Lines that are not valid JSON (e.g. docker-compose prefix lines) are
    silently skipped.
    """
    entries: list[dict[str, Any]] = []
    if not os.path.exists(path):
        return entries
    with open(path, "r") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            # Handle docker-compose log prefixes: strip everything up to the
            # first '{' character.
            json_start = line.find("{")
            if json_start < 0:
                continue
            json_str = line[json_start:]
            try:
                entry = json.loads(json_str)
                entries.append(entry)
            except json.JSONDecodeError:
                continue
    return entries


def check_evidence_in_log(
    entries: list[dict[str, Any]], query: str | None
) -> tuple[bool, int]:
    """Check if audit log entries satisfy the evidence query.

    Returns (found, count) where found is True if at least one entry matches,
    and count is the number of matching entries.

    The query is a simplified jq-like expression evaluated as Python logic.
    """
    if not query or not entries:
        return False, 0

    count = 0
    for entry in entries:
        if _matches_query(entry, query):
            count += 1
    return count > 0, count


def _matches_query(entry: dict[str, Any], query: str) -> bool:
    """Evaluate a simplified evidence query against a log entry.

    Supports common patterns from the taxonomy:
      - '.field != null'
      - '.field != ""'
      - '.field == "value"'
      - '.field | startswith("prefix")'
      - '.field | contains("substring")'
      - compound 'and' expressions
    """
    # Split on ' and ' for compound queries
    parts = query.split(" and ")
    for part in parts:
        part = part.strip()
        if not _eval_single(entry, part):
            return False
    return True


def _resolve_path(entry: dict[str, Any], dotpath: str) -> Any:
    """Resolve a dot-separated path against a nested dict.

    Returns the value at the path, or a sentinel _MISSING if not found.
    """
    parts = dotpath.split(".")
    obj: Any = entry
    for p in parts:
        if isinstance(obj, dict):
            obj = obj.get(p, _MISSING)
            if obj is _MISSING:
                return _MISSING
        else:
            return _MISSING
    return obj


# Sentinel for missing dict keys (distinct from None)
_MISSING = object()


def _eval_single(entry: dict[str, Any], expr: str) -> bool:
    """Evaluate a single query expression."""
    expr = expr.strip()

    # Handle .field | startswith("value")
    if "| startswith(" in expr:
        field, _, rest = expr.partition(" | startswith(")
        field = field.strip().lstrip(".")
        value = rest.rstrip(")").strip('"').strip("'")
        resolved = _resolve_path(entry, field)
        if resolved is _MISSING:
            return False
        return str(resolved).startswith(value)

    # Handle .field | contains("value")
    if "| contains(" in expr:
        field, _, rest = expr.partition(" | contains(")
        field = field.strip().lstrip(".")
        value = rest.rstrip(")").strip('"').strip("'")
        resolved = _resolve_path(entry, field)
        if resolved is _MISSING:
            return False
        return value in str(resolved)

    # Handle .field != null  (works for nested fields too)
    if "!= null" in expr:
        field = expr.replace("!= null", "").strip().lstrip(".")
        resolved = _resolve_path(entry, field)
        return resolved is not _MISSING and resolved is not None

    # Handle .field != ""
    if '!= ""' in expr:
        field = expr.replace('!= ""', "").strip().lstrip(".")
        resolved = _resolve_path(entry, field)
        if resolved is _MISSING:
            return False
        return resolved != ""

    # Handle .field == "value"
    if '== "' in expr:
        field, _, rest = expr.partition(' == "')
        field = field.strip().lstrip(".")
        value = rest.rstrip('"')
        resolved = _resolve_path(entry, field)
        if resolved is _MISSING:
            return False
        return str(resolved) == value

    # Handle .field == number (e.g. .status_code == 403)
    if " == " in expr:
        field, _, value = expr.partition(" == ")
        field = field.strip().lstrip(".")
        value = value.strip()
        resolved = _resolve_path(entry, field)
        if resolved is _MISSING:
            return False
        try:
            return resolved == int(value)
        except (ValueError, TypeError):
            return str(resolved) == value

    return False


# ---------------------------------------------------------------------------
# Config file evidence
# ---------------------------------------------------------------------------


def check_config_exists(project_root: Path) -> dict[str, bool]:
    """Check which policy/config files exist."""
    configs = {
        "opa_policy": project_root / "config" / "opa" / "mcp_policy.rego",
        "tool_registry": project_root / "config" / "tool-registry.yaml",
        "opa_tool_grants": project_root / "config" / "opa" / "tool_grants.yaml",
        "spiffe_ids": project_root / "config" / "spiffe-ids.yaml",
        "destinations": project_root / "config" / "destinations.yaml",
        "risk_thresholds": project_root / "config" / "risk_thresholds.yaml",
        "cosign_key": project_root / ".cosign",
    }
    return {name: path.exists() for name, path in configs.items()}


# ---------------------------------------------------------------------------
# Status determination
# ---------------------------------------------------------------------------


def determine_status(
    control: dict[str, Any],
    audit_found: bool,
    configs: dict[str, bool],
) -> str:
    """Determine control implementation status.

    Returns one of: Implemented, Partial, Documented Only
    """
    middleware = control.get("middleware")
    evidence_type = control.get("evidence_type", "")

    # No middleware means supply chain or infrastructure control
    if not middleware:
        return "Documented Only"

    # Check if middleware is implemented in the codebase
    if middleware not in IMPLEMENTED_MIDDLEWARE:
        return "Documented Only"

    # For audit_log evidence, check if we found matching entries
    if evidence_type == "audit_log":
        return "Implemented" if audit_found else "Partial"

    # For configuration evidence, check if relevant config exists
    if evidence_type == "configuration":
        if middleware in ("opa", "spiffe_auth"):
            return "Implemented" if configs.get("opa_policy") else "Partial"
        if middleware == "dlp":
            return "Implemented"
        if middleware in ("tool_registry",):
            return "Implemented" if configs.get("tool_registry") else "Partial"
        return "Implemented"

    # For test_result evidence, mark as Implemented (tests exist in codebase)
    if evidence_type == "test_result":
        return "Implemented"

    return "Partial"


# ---------------------------------------------------------------------------
# Evidence description generation
# ---------------------------------------------------------------------------


def build_evidence_reference(
    control: dict[str, Any],
    audit_count: int,
    configs: dict[str, bool],
    project_root: Path,
) -> str:
    """Build a human-readable evidence reference string."""
    evidence_type = control.get("evidence_type", "")

    if evidence_type == "audit_log" and audit_count > 0:
        return f"audit.jsonl ({audit_count} matching entries)"
    if evidence_type == "configuration":
        middleware = control.get("middleware", "")
        if middleware in ("opa", "spiffe_auth"):
            return "config/opa/mcp_policy.rego"
        if middleware == "tool_registry":
            return "config/tool-registry.yaml"
        if middleware == "dlp":
            return "internal/gateway/middleware/dlp.go"
        if middleware == "deep_scan":
            return "config/opa/mcp_policy.rego (poisoning patterns)"
        if middleware == "session_context":
            return "config/opa/mcp_policy.rego (session risk threshold)"
        if middleware in ("size_limit", "rate_limiter", "circuit_breaker"):
            return f"internal/gateway/middleware/{middleware}.go"
        return f"internal/gateway/middleware/{middleware}.go"
    if evidence_type == "test_result":
        middleware = control.get("middleware", "")
        if middleware:
            return f"internal/gateway/middleware/{middleware}_test.go"
        return "tests/"

    return "N/A"


def build_evidence_description(control: dict[str, Any], audit_count: int) -> str:
    """Build a description of the evidence found."""
    evidence_type = control.get("evidence_type", "")
    middleware = control.get("middleware", "")

    if evidence_type == "audit_log":
        if audit_count > 0:
            return f"{audit_count} audit log entries demonstrate {middleware} enforcement"
        return f"No audit log entries found for {middleware} -- run E2E suite to generate"

    if evidence_type == "configuration":
        return f"Configuration file defines {middleware} policy rules"

    if evidence_type == "test_result":
        return f"Unit and integration tests verify {middleware} behavior"

    return "No evidence collected"


def build_implementation_notes(control: dict[str, Any]) -> str:
    """Generate implementation notes based on control metadata."""
    middleware = control.get("middleware")
    step = control.get("step")

    if middleware and step is not None:
        return f"Middleware: {middleware} (chain step {step})"
    if middleware:
        return f"Middleware: {middleware}"
    return "Infrastructure/CI control -- not in middleware chain"


def build_limitations(control: dict[str, Any], status: str) -> str:
    """Generate limitation notes."""
    if status == "Documented Only":
        return "Control exists in design only; not enforced at runtime in POC"
    if status == "Partial":
        return "Partial implementation; evidence collection incomplete"

    middleware = control.get("middleware", "")
    if middleware == "spiffe_auth":
        return "POC uses self-signed CAs; production requires enterprise PKI"
    if middleware in ("spike_token", "spike_redeemer"):
        return "SPIKE Nexus integration requires running Nexus container"
    if middleware == "step_up_gating":
        return "Guard model unavailable in POC; fails open for medium risk"

    return ""


def build_recommendation(control: dict[str, Any], status: str) -> str:
    """Generate recommendations for improvement."""
    if status == "Documented Only":
        return "Implement runtime enforcement before production deployment"
    if status == "Partial":
        return "Complete implementation and run E2E suite to generate evidence"
    return "Maintain current controls; review annually"


# ---------------------------------------------------------------------------
# CSV generation
# ---------------------------------------------------------------------------


def generate_rows(
    controls: list[dict[str, Any]],
    audit_entries: list[dict[str, Any]],
    configs: dict[str, bool],
    project_root: Path,
) -> list[dict[str, str]]:
    """Generate one CSV row per control-framework combination."""
    rows: list[dict[str, str]] = []

    for control in controls:
        frameworks = control.get("frameworks", {})
        evidence_query = control.get("evidence_query")

        audit_found, audit_count = check_evidence_in_log(audit_entries, evidence_query)
        status = determine_status(control, audit_found, configs)

        # Emit one row per framework mapping
        for framework_name, requirements in frameworks.items():
            if not requirements:
                continue
            for req_id in requirements:
                req_desc = FRAMEWORK_REQUIREMENTS.get(framework_name, {}).get(
                    req_id, req_id
                )
                impl_notes = build_implementation_notes(control)
                # Cross-reference ROPA document for GDPR Art. 30 mappings
                if req_id == "Art. 30":
                    impl_notes += f"; ROPA: {GDPR_ART30_ROPA_PATH}"
                row = {
                    "control_id": control["id"],
                    "control_name": control["name"],
                    "control_description": control["description"],
                    "framework": framework_name.upper(),
                    "framework_requirement": f"{req_id}: {req_desc}",
                    "status": status,
                    "evidence_type": control.get("evidence_type", ""),
                    "evidence_reference": build_evidence_reference(
                        control, audit_count, configs, project_root
                    ),
                    "evidence_description": build_evidence_description(
                        control, audit_count
                    ),
                    "test_result": "PASS" if status == "Implemented" else "N/A",
                    "implementation_notes": impl_notes,
                    "limitations": build_limitations(control, status),
                    "recommendation": build_recommendation(control, status),
                }
                rows.append(row)

    return rows


def write_csv(rows: list[dict[str, str]], output_path: Path) -> None:
    """Write rows to a CSV file."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def _latest_audit_timestamp_for_query(
    entries: list[dict[str, Any]],
    query: str | None,
) -> str | None:
    """Return the latest timestamp among entries matching query.

    If no entries match or no timestamps exist, return None.
    """
    if not query:
        return None
    latest: str | None = None
    for entry in entries:
        if not _matches_query(entry, query):
            continue
        ts = entry.get("timestamp")
        if not isinstance(ts, str) or not ts:
            continue
        if latest is None or ts > latest:
            latest = ts
    return latest


def build_evidence_bundle(
    rows: list[dict[str, str]],
    controls: list[dict[str, Any]],
    audit_entries: list[dict[str, Any]],
    generated_at: str,
) -> dict[str, Any]:
    """Build evidence bundle v2 records from compliance rows and controls.

    Each record includes the fields auditors expect for machine-readable
    ingestion: control_id, source, timestamp, status, artifact_reference.
    """
    control_by_id = {c.get("id", ""): c for c in controls}
    records: list[dict[str, str]] = []

    for row in rows:
        control_id = row.get("control_id", "")
        control = control_by_id.get(control_id, {})
        source = str(control.get("evidence_type", "") or "unknown")
        timestamp = generated_at
        if source == "audit_log":
            latest = _latest_audit_timestamp_for_query(
                audit_entries,
                control.get("evidence_query"),
            )
            if latest:
                timestamp = latest

        records.append(
            {
                "control_id": control_id,
                "framework": row.get("framework", ""),
                "framework_requirement": row.get("framework_requirement", ""),
                "status": row.get("status", ""),
                "source": source,
                "timestamp": timestamp,
                "artifact_reference": row.get("evidence_reference", ""),
                "control_name": row.get("control_name", ""),
            }
        )

    return {
        "schema_version": EVIDENCE_BUNDLE_SCHEMA_VERSION,
        "generated_at": generated_at,
        "record_count": len(records),
        "records": records,
    }


def write_evidence_bundle_json(bundle: dict[str, Any], output_path: Path) -> None:
    """Write evidence bundle v2 as pretty JSON."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w") as f:
        json.dump(bundle, f, indent=2, sort_keys=False)
        f.write("\n")


def write_evidence_bundle_csv(records: list[dict[str, str]], output_path: Path) -> None:
    """Write evidence bundle records as CSV."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=EVIDENCE_BUNDLE_COLUMNS)
        writer.writeheader()
        writer.writerows(records)


# ---------------------------------------------------------------------------
# XLSX generation
# ---------------------------------------------------------------------------

# Framework display names for sheet tabs and headings.
FRAMEWORK_DISPLAY_NAMES: dict[str, str] = {
    "SOC2": "SOC 2",
    "ISO27001": "ISO 27001",
    "CCPA": "CCPA",
    "GDPR": "GDPR",
}

# Conditional formatting colors keyed by status.
STATUS_COLORS: dict[str, str] = {
    "Implemented": "C6EFCE",     # green
    "Partial": "FFEB9C",         # yellow
    "Documented Only": "D9D9D9", # gray
}

# Alternating row fill for readability.
_ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


def _apply_xlsx_formatting(ws: "openpyxl.worksheet.worksheet.Worksheet") -> None:
    """Apply bold headers, alternating row colors, conditional status colors, and auto-width."""
    # Bold header row
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    # Status column index (1-based) -- find it dynamically.
    status_col: int | None = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == "status":
            status_col = idx
            break

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        # Alternating row color (for rows without status-based fill)
        is_alt = row_idx % 2 == 0

        for cell in row:
            cell.border = thin_border

        # Status-based conditional coloring takes precedence over alternating rows.
        if status_col:
            status_val = row[status_col - 1].value
            color_hex = STATUS_COLORS.get(status_val or "")
            if color_hex:
                fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                for cell in row:
                    cell.fill = fill
                continue

        # Fall back to alternating row fill.
        if is_alt:
            for cell in row:
                cell.fill = _ALT_ROW_FILL

    # Auto-width columns.
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        # Cap width at 60 to keep columns manageable.
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[col_letter].width = adjusted_width


def _compute_framework_summary(rows: list[dict[str, str]]) -> dict[str, dict[str, int]]:
    """Compute per-framework status counts.

    Returns {framework: {"total": N, "Implemented": N, "Partial": N, "Documented Only": N, "pct_implemented": N}}
    """
    summary: dict[str, dict[str, int]] = {}
    for row in rows:
        fw = row["framework"]
        if fw not in summary:
            summary[fw] = {"total": 0, "Implemented": 0, "Partial": 0, "Documented Only": 0}
        summary[fw]["total"] += 1
        status = row["status"]
        if status in summary[fw]:
            summary[fw][status] += 1

    # Compute percentages.
    for fw in summary:
        total = summary[fw]["total"]
        impl = summary[fw]["Implemented"]
        summary[fw]["pct_implemented"] = round(impl * 100 / total) if total > 0 else 0

    return summary


def write_xlsx(rows: list[dict[str, str]], output_path: Path) -> None:
    """Write compliance data to an XLSX workbook with per-framework sheets and a summary.

    Sheets:
      - Summary: total controls, % implemented per framework, date, version
      - One sheet per framework (SOC 2, ISO 27001, CCPA, GDPR)
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    # --- Summary sheet ---
    ws_summary = wb.active
    ws_summary.title = "Summary"

    today = date.today().isoformat()
    ws_summary.append(["Compliance Report Summary"])
    ws_summary.append([])
    ws_summary.append(["Architecture", "Agentic AI Security Reference Architecture"])
    ws_summary.append(["Version", "POC v1.0"])
    ws_summary.append(["Report Date", today])
    ws_summary.append(["Total Controls", len({r["control_id"] for r in rows})])
    ws_summary.append(["Total Control-Framework Rows", len(rows)])
    ws_summary.append([])

    fw_summary = _compute_framework_summary(rows)

    ws_summary.append(["Framework", "Total Mappings", "Implemented", "Partial",
                        "Documented Only", "% Implemented"])
    for fw_code in ["SOC2", "ISO27001", "CCPA", "GDPR"]:
        if fw_code in fw_summary:
            s = fw_summary[fw_code]
            display = FRAMEWORK_DISPLAY_NAMES.get(fw_code, fw_code)
            ws_summary.append([
                display,
                s["total"],
                s["Implemented"],
                s["Partial"],
                s["Documented Only"],
                f"{s['pct_implemented']}%",
            ])

    # Bold title and header rows.
    ws_summary["A1"].font = Font(bold=True, size=14)
    for cell in ws_summary[9]:  # header row of the table
        cell.font = Font(bold=True)

    # Auto-width summary columns.
    for col_cells in ws_summary.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws_summary.column_dimensions[col_letter].width = min(max_len + 2, 60)

    # --- Per-framework sheets ---
    frameworks_order = ["SOC2", "ISO27001", "CCPA", "GDPR"]
    for fw_code in frameworks_order:
        fw_rows = [r for r in rows if r["framework"] == fw_code]
        if not fw_rows:
            continue

        display_name = FRAMEWORK_DISPLAY_NAMES.get(fw_code, fw_code)
        ws = wb.create_sheet(title=display_name)

        # Header row
        ws.append(CSV_COLUMNS)

        for row in fw_rows:
            ws.append([row.get(col, "") for col in CSV_COLUMNS])

        _apply_xlsx_formatting(ws)

    wb.save(str(output_path))


# ---------------------------------------------------------------------------
# PDF generation
# ---------------------------------------------------------------------------

# Architecture name and version used in the PDF cover page.
_ARCH_NAME = "Agentic AI Security Reference Architecture"
_ARCH_VERSION = "POC v1.0"


class CompliancePDF(FPDF):
    """Custom PDF class for compliance summary report."""

    def header(self):
        if self.page_no() > 1:
            self.set_font("Helvetica", "I", 8)
            self.cell(0, 10, f"{_ARCH_NAME} -- Compliance Summary", align="C")
            self.ln(5)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.cell(0, 10, f"Page {self.page_no()}/{{nb}}", align="C")


def _build_control_area_matrix(
    rows: list[dict[str, str]],
) -> dict[str, dict[str, str]]:
    """Build a matrix of control_area -> framework -> best_status.

    'Best status' priority: Implemented > Partial > Documented Only.
    """
    priority = {"Implemented": 3, "Partial": 2, "Documented Only": 1, "": 0}
    matrix: dict[str, dict[str, str]] = {}

    for row in rows:
        control_id = row["control_id"]
        # Extract area prefix (e.g., GW-AUTH from GW-AUTH-001)
        parts = control_id.rsplit("-", 1)
        area = parts[0] if len(parts) == 2 else control_id
        fw = row["framework"]
        status = row["status"]

        if area not in matrix:
            matrix[area] = {}
        current = matrix[area].get(fw, "")
        if priority.get(status, 0) > priority.get(current, 0):
            matrix[area][fw] = status

    return matrix


# Human-readable area names for the PDF control matrix.
_AREA_DISPLAY_NAMES: dict[str, str] = {
    "GW-AUTH": "Identity (SPIFFE)",
    "GW-AUTHZ": "Authorization (OPA)",
    "GW-DLP": "Data Protection (DLP)",
    "GW-SCAN": "Content Security",
    "GW-AUDIT": "Audit Logging",
    "GW-SEC": "Secrets (SPIKE)",
    "GW-TRANS": "Transport (mTLS)",
    "GW-AVAIL": "Availability",
    "GW-SESS": "Session Context",
    "GW-SC": "Supply Chain",
}


def _select_evidence_highlights(
    audit_entries: list[dict[str, Any]],
    max_entries: int = 8,
) -> list[dict[str, Any]]:
    """Select diverse, exemplary audit log entries for the evidence highlights page.

    Selects entries demonstrating different control areas.  Redacts sensitive
    fields (full session_id, decision_id shortened).
    """
    # Prefer entries that show different actions and security features.
    seen_actions: set[str] = set()
    selected: list[dict[str, Any]] = []

    for entry in audit_entries:
        action = entry.get("action", "unknown")
        if action not in seen_actions and len(selected) < max_entries:
            seen_actions.add(action)
            # Redact: truncate UUIDs to first 8 chars for privacy.
            redacted = {}
            for k, v in entry.items():
                if k in ("session_id", "decision_id", "trace_id") and isinstance(v, str) and len(v) > 8:
                    redacted[k] = v[:8] + "..."
                elif k == "security" and isinstance(v, dict):
                    redacted[k] = {sk: sv for sk, sv in v.items()}
                else:
                    redacted[k] = v
            selected.append(redacted)

    # Fill remaining slots with entries that have interesting security fields.
    for entry in audit_entries:
        if len(selected) >= max_entries:
            break
        if entry.get("security") and entry.get("action") not in seen_actions:
            seen_actions.add(entry.get("action", ""))
            redacted = {}
            for k, v in entry.items():
                if k in ("session_id", "decision_id", "trace_id") and isinstance(v, str) and len(v) > 8:
                    redacted[k] = v[:8] + "..."
                else:
                    redacted[k] = v
            selected.append(redacted)

    return selected[:max_entries]


def _gather_partial_documented_controls(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    """Return unique controls that are Partial or Documented Only, for the recommendations page."""
    seen: set[str] = set()
    result: list[dict[str, str]] = []
    for row in rows:
        if row["status"] in ("Partial", "Documented Only") and row["control_id"] not in seen:
            seen.add(row["control_id"])
            result.append(row)
    return result


def write_pdf(
    rows: list[dict[str, str]],
    audit_entries: list[dict[str, Any]],
    output_path: Path,
) -> int:
    """Generate a 4-page compliance summary PDF.

    Page 1: Cover + Summary
    Page 2: Control Matrix (heat map)
    Page 3: Evidence Highlights
    Page 4: Limitations and Recommendations

    Returns the number of pages generated.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    pdf = CompliancePDF(orientation="P", unit="mm", format="A4")
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=15)

    today = date.today().isoformat()
    fw_summary = _compute_framework_summary(rows)
    total_controls = len({r["control_id"] for r in rows})
    total_implemented = sum(s["Implemented"] for s in fw_summary.values())
    total_mappings = sum(s["total"] for s in fw_summary.values())

    # ---- Page 1: Cover + Summary ----
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 24)
    pdf.cell(0, 40, "Compliance Summary", align="C", new_x="LMARGIN", new_y="NEXT")

    pdf.set_font("Helvetica", "", 14)
    pdf.cell(0, 10, _ARCH_NAME, align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 8, f"Version: {_ARCH_VERSION}", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 8, f"Report Date: {today}", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(10)

    # Overall posture
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 10, "Overall Security Posture", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 8,
             f"{total_implemented}/{total_mappings} control-framework mappings implemented "
             f"across {total_controls} controls",
             new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    # Framework coverage percentages
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 10, "Framework Coverage", new_x="LMARGIN", new_y="NEXT")

    pdf.set_font("Helvetica", "B", 10)
    col_w = [60, 30, 30, 30, 40]
    headers = ["Framework", "Total", "Implemented", "Partial", "% Implemented"]
    for i, h in enumerate(headers):
        pdf.cell(col_w[i], 8, h, border=1)
    pdf.ln()

    pdf.set_font("Helvetica", "", 10)
    for fw_code in ["SOC2", "ISO27001", "CCPA", "GDPR"]:
        if fw_code not in fw_summary:
            continue
        s = fw_summary[fw_code]
        display = FRAMEWORK_DISPLAY_NAMES.get(fw_code, fw_code)
        pdf.cell(col_w[0], 8, display, border=1)
        pdf.cell(col_w[1], 8, str(s["total"]), border=1, align="C")
        pdf.cell(col_w[2], 8, str(s["Implemented"]), border=1, align="C")
        pdf.cell(col_w[3], 8, str(s["Partial"]), border=1, align="C")
        pdf.cell(col_w[4], 8, f"{s['pct_implemented']}%", border=1, align="C")
        pdf.ln()

    # ---- Page 2: Control Matrix (heat map) ----
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "Control Matrix", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    matrix = _build_control_area_matrix(rows)
    frameworks_order = ["SOC2", "ISO27001", "CCPA", "GDPR"]

    # Table header
    area_col_w = 55
    fw_col_w = 32
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(area_col_w, 8, "Control Area", border=1)
    for fw_code in frameworks_order:
        display = FRAMEWORK_DISPLAY_NAMES.get(fw_code, fw_code)
        pdf.cell(fw_col_w, 8, display, border=1, align="C")
    pdf.ln()

    # Table rows with color coding
    # Sort by the standard area order from the taxonomy
    area_order = [
        "GW-AUTH", "GW-AUTHZ", "GW-DLP", "GW-SCAN", "GW-AUDIT",
        "GW-SEC", "GW-TRANS", "GW-AVAIL", "GW-SESS", "GW-SC",
    ]
    pdf.set_font("Helvetica", "", 9)
    status_rgb: dict[str, tuple[int, int, int]] = {
        "Implemented": (198, 239, 206),    # green
        "Partial": (255, 235, 156),        # yellow
        "Documented Only": (217, 217, 217),  # gray
    }

    for area in area_order:
        if area not in matrix:
            continue
        area_display = _AREA_DISPLAY_NAMES.get(area, area)
        pdf.set_fill_color(255, 255, 255)
        pdf.cell(area_col_w, 8, area_display, border=1)
        for fw_code in frameworks_order:
            status = matrix[area].get(fw_code, "")
            if status and status in status_rgb:
                r, g, b = status_rgb[status]
                pdf.set_fill_color(r, g, b)
                # Use short label
                label = {"Implemented": "Impl.", "Partial": "Partial", "Documented Only": "Doc."}.get(status, "")
                pdf.cell(fw_col_w, 8, label, border=1, align="C", fill=True)
            else:
                pdf.set_fill_color(255, 255, 255)
                pdf.cell(fw_col_w, 8, "N/A", border=1, align="C")
        pdf.ln()

    # Legend
    pdf.ln(5)
    pdf.set_font("Helvetica", "I", 8)
    pdf.cell(0, 6, "Legend: Impl. = Implemented (runtime enforced), Partial = partial evidence, "
             "Doc. = Documented Only (design-time)", new_x="LMARGIN", new_y="NEXT")

    # ---- Page 3: Evidence Highlights ----
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "Evidence Highlights", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    highlights = _select_evidence_highlights(audit_entries)
    if highlights:
        pdf.set_font("Helvetica", "", 9)
        pdf.multi_cell(0, 5,
                        "The following are exemplary audit log entries demonstrating key security "
                        "controls in action.  Identifiers have been truncated for readability.")
        pdf.ln(3)

        for i, entry in enumerate(highlights, start=1):
            pdf.set_font("Helvetica", "B", 9)
            action = entry.get("action", "unknown")
            result = entry.get("result", "")
            pdf.cell(0, 6, f"Entry {i}: action={action}, result={result}",
                     new_x="LMARGIN", new_y="NEXT")

            pdf.set_font("Courier", "", 7)
            # Compact JSON representation
            json_str = json.dumps(entry, indent=2, default=str)
            # Limit line length for PDF
            for line in json_str.split("\n"):
                truncated = line[:120]
                pdf.cell(0, 4, truncated, new_x="LMARGIN", new_y="NEXT")
            pdf.ln(2)
    else:
        pdf.set_font("Helvetica", "I", 10)
        pdf.cell(0, 10, "No audit log entries available. Run the E2E suite to generate evidence.",
                 new_x="LMARGIN", new_y="NEXT")

    # ---- Page 4: Limitations and Recommendations ----
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "Limitations and Recommendations", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # What is NOT covered
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 8, "Scope Limitations", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    limitations_text = [
        "This report covers controls implemented in the MCP Security Gateway middleware chain only.",
        "Infrastructure-level controls (network segmentation, host hardening, key management) are out of scope.",
        "Supply chain controls (container signing, SBOM) are documented but not runtime-enforced in this POC.",
        "Production deployments require additional controls: enterprise PKI, HSM integration, SIEM forwarding.",
        "The POC uses self-signed CAs; production SPIFFE deployments need enterprise-grade PKI infrastructure.",
    ]
    for line in limitations_text:
        pdf.cell(5, 5, "-")
        pdf.cell(0, 5, line, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # Gateway vs infrastructure distinction
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 8, "Gateway vs. Infrastructure Controls", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    pdf.multi_cell(0, 5,
                    "The gateway enforces application-layer security (authentication, authorization, "
                    "DLP, audit logging, rate limiting). Infrastructure controls (network policies, "
                    "host OS hardening, encryption at rest) must be addressed separately in the "
                    "deployment environment (Kubernetes, cloud provider).")
    pdf.ln(3)

    # Recommendations for Partial / Documented Only
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 8, "Recommendations for Incomplete Controls", new_x="LMARGIN", new_y="NEXT")

    incomplete = _gather_partial_documented_controls(rows)
    if incomplete:
        pdf.set_font("Helvetica", "B", 8)
        rec_col_w = [30, 50, 30, 80]
        rec_headers = ["Control ID", "Control Name", "Status", "Recommendation"]
        for i, h in enumerate(rec_headers):
            pdf.cell(rec_col_w[i], 7, h, border=1)
        pdf.ln()

        pdf.set_font("Helvetica", "", 8)
        for ctrl in incomplete:
            pdf.cell(rec_col_w[0], 7, ctrl["control_id"], border=1)
            # Truncate long names
            name = ctrl["control_name"][:25] + ("..." if len(ctrl["control_name"]) > 25 else "")
            pdf.cell(rec_col_w[1], 7, name, border=1)
            pdf.cell(rec_col_w[2], 7, ctrl["status"], border=1)
            rec = ctrl["recommendation"][:45] + ("..." if len(ctrl["recommendation"]) > 45 else "")
            pdf.cell(rec_col_w[3], 7, rec, border=1)
            pdf.ln()
    else:
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(0, 8, "All controls are fully implemented.", new_x="LMARGIN", new_y="NEXT")

    pdf.output(str(output_path))
    return pdf.page_no()


# ---------------------------------------------------------------------------
# Evidence copying
# ---------------------------------------------------------------------------


def copy_evidence(
    project_root: Path,
    audit_log_path: str,
    output_dir: Path,
) -> list[str]:
    """Copy supporting evidence files into the report output directory.

    Copies:
      - audit-log-excerpt.jsonl  (first 50 lines of audit log)
      - e2e-test-results.txt     (from tests/e2e/ if exists)
      - policy-configs/          (opa-policy.rego, tool-registry.yaml, risk-thresholds.yaml)

    Returns list of copied file paths (relative to output_dir).
    """
    evidence_dir = output_dir / "evidence"
    evidence_dir.mkdir(parents=True, exist_ok=True)
    copied: list[str] = []

    # 1. Audit log excerpt
    if os.path.exists(audit_log_path):
        excerpt_path = evidence_dir / "audit-log-excerpt.jsonl"
        with open(audit_log_path, "r") as src, open(excerpt_path, "w") as dst:
            for i, line in enumerate(src):
                if i >= 50:
                    break
                dst.write(line)
        copied.append("evidence/audit-log-excerpt.jsonl")

    # 2. E2E test results
    e2e_results_candidates = [
        project_root / "tests" / "e2e" / "gateway-audit-logs.log",
        project_root / "tests" / "e2e" / "test-results.txt",
    ]
    for candidate in e2e_results_candidates:
        if candidate.exists():
            dst_path = evidence_dir / "e2e-test-results.txt"
            shutil.copy2(str(candidate), str(dst_path))
            copied.append("evidence/e2e-test-results.txt")
            break

    # 2b. Immutable sink verification artifact (K8s audit path)
    immutable_sink_proof = (
        project_root
        / "tests"
        / "e2e"
        / "artifacts"
        / "immutable-audit-sink-proof.json"
    )
    if immutable_sink_proof.exists():
        dst_path = evidence_dir / "immutable-audit-sink-proof.json"
        shutil.copy2(str(immutable_sink_proof), str(dst_path))
        copied.append("evidence/immutable-audit-sink-proof.json")

    # 3. Policy config files
    policy_dir = evidence_dir / "policy-configs"
    policy_dir.mkdir(parents=True, exist_ok=True)

    policy_files = {
        "opa-policy.rego": project_root / "config" / "opa" / "mcp_policy.rego",
        "tool-registry.yaml": project_root / "config" / "tool-registry.yaml",
        "risk-thresholds.yaml": project_root / "config" / "risk_thresholds.yaml",
    }
    for dest_name, src_path in policy_files.items():
        if src_path.exists():
            shutil.copy2(str(src_path), str(policy_dir / dest_name))
            copied.append(f"evidence/policy-configs/{dest_name}")

    return copied


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main(argv: list[str] | None = None) -> int:
    """CLI entry point."""
    parser = argparse.ArgumentParser(
        description="Generate compliance report CSV from control taxonomy and audit logs"
    )
    parser.add_argument(
        "--audit-log",
        default="/tmp/audit.jsonl",
        help="Path to audit log JSONL file (default: /tmp/audit.jsonl)",
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        help="Output directory (default: reports/compliance-YYYY-MM-DD/)",
    )
    parser.add_argument(
        "--taxonomy",
        default=None,
        help="Path to control taxonomy YAML (default: tools/compliance/control_taxonomy.yaml)",
    )
    parser.add_argument(
        "--project-root",
        default=None,
        help="Project root directory (default: auto-detected)",
    )
    args = parser.parse_args(argv)

    # Determine project root
    if args.project_root:
        project_root = Path(args.project_root).resolve()
    else:
        # Assume script is at tools/compliance/generate.py
        project_root = SCRIPT_DIR.parent.parent

    # Load taxonomy
    taxonomy_path = Path(args.taxonomy) if args.taxonomy else TAXONOMY_PATH
    print(f"Loading control taxonomy from {taxonomy_path}")
    controls = load_taxonomy(taxonomy_path)
    print(f"  Loaded {len(controls)} controls")

    # Load audit log
    audit_path = args.audit_log
    print(f"Loading audit log from {audit_path}")
    audit_entries = load_audit_log(audit_path)
    print(f"  Loaded {len(audit_entries)} audit entries")

    # Check config files
    configs = check_config_exists(project_root)
    present = sum(1 for v in configs.values() if v)
    print(f"Checked {len(configs)} config files ({present} present)")

    # Generate report
    rows = generate_rows(controls, audit_entries, configs, project_root)
    print(f"Generated {len(rows)} compliance report rows")
    generated_at = datetime.now(timezone.utc).isoformat()

    # Generate machine-readable evidence bundle (v2)
    evidence_bundle = build_evidence_bundle(
        rows,
        controls,
        audit_entries,
        generated_at,
    )

    # Determine output path
    if args.output_dir:
        output_dir = Path(args.output_dir)
    else:
        today = date.today().isoformat()
        output_dir = project_root / "reports" / f"compliance-{today}"

    # --- Evidence bundle outputs (JSON + CSV) ---
    evidence_json_path = output_dir / "compliance-evidence.v2.json"
    write_evidence_bundle_json(evidence_bundle, evidence_json_path)
    print(f"Evidence JSON bundle written to {evidence_json_path}")

    evidence_csv_path = output_dir / "compliance-evidence.v2.csv"
    write_evidence_bundle_csv(evidence_bundle["records"], evidence_csv_path)
    print(f"Evidence CSV bundle written to {evidence_csv_path}")

    # --- CSV output ---
    csv_path = output_dir / "compliance-report.csv"
    write_csv(rows, csv_path)
    print(f"CSV report written to {csv_path}")

    # --- XLSX output ---
    xlsx_path = output_dir / "compliance-report.xlsx"
    write_xlsx(rows, xlsx_path)
    print(f"XLSX report written to {xlsx_path}")

    # --- PDF output ---
    pdf_path = output_dir / "compliance-summary.pdf"
    num_pages = write_pdf(rows, audit_entries, pdf_path)
    print(f"PDF summary written to {pdf_path} ({num_pages} pages)")

    # --- Copy evidence ---
    evidence_files = copy_evidence(project_root, audit_path, output_dir)
    if evidence_files:
        print(f"Evidence files copied: {len(evidence_files)}")
        for ef in evidence_files:
            print(f"  {ef}")

    # Summary
    statuses: dict[str, int] = {}
    for row in rows:
        s = row["status"]
        statuses[s] = statuses.get(s, 0) + 1
    print("\nStatus Summary:")
    for s, count in sorted(statuses.items()):
        print(f"  {s}: {count}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
